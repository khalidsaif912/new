import hashlib
import json
import os
import re
import time
from io import BytesIO
from urllib.parse import parse_qsl, urlencode, urlparse, urlunparse

import requests
from openpyxl import load_workbook

from roster_app.settings import ROSTERS_DIR, SOURCE_NAME_FALLBACK, SOURCE_NAME_URL

DEBUG_SHAREPOINT_RESPONSE_PATH = "debug_sharepoint_response.png"


def _add_or_replace_query_param(url: str, key: str, value: str) -> str:
    u = urlparse(url)
    qs = dict(parse_qsl(u.query, keep_blank_values=True))
    qs[key] = value
    return urlunparse(u._replace(query=urlencode(qs, doseq=True)))


def _normalize_sharepoint_download_url(url: str, *, cache_bust: bool = True) -> str:
    if not url:
        return url
    u = urlparse(url)
    host = (u.netloc or "").lower()
    if ("sharepoint.com" not in host) and ("onedrive.live.com" not in host) and ("1drv.ms" not in host):
        out = url
    else:
        out = _add_or_replace_query_param(url, "download", "1")
        out = _add_or_replace_query_param(out, "web", "0")
    # Bust CDN/proxy caches when the sharing link is reused for an overwritten file.
    if cache_bust:
        out = _add_or_replace_query_param(out, "_cb", str(int(time.time() * 1000)))
    return out


def workbook_content_fingerprint(data: bytes) -> str:
    """
    Logical fingerprint of sheet cell values (ignores ZIP/xlsx metadata noise).
    Used to detect same-filename overwrites even when CDN/metadata quirks occur.
    """
    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    h = hashlib.sha256()
    try:
        for sheet_name in wb.sheetnames:
            h.update(sheet_name.encode("utf-8", errors="replace"))
            h.update(b"\0")
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell is None:
                        h.update(b"|")
                    else:
                        h.update(str(cell).encode("utf-8", errors="replace"))
                        h.update(b"|")
                h.update(b"\n")
    finally:
        wb.close()
    return h.hexdigest()


def _file_signature_hex16(data: bytes) -> str:
    return (data[:16] or b"").hex()


def _is_excel_signature(data: bytes) -> bool:
    head = data[:8] or b""
    # xlsx/zip
    if data.startswith(b"PK\x03\x04"):
        return True
    # xls (OLE compound)
    if head.startswith(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"):
        return True
    return False


def _is_png_signature(data: bytes) -> bool:
    return (data[:8] or b"").startswith(b"\x89PNG\r\n\x1a\n")


def download_excel(url: str) -> bytes:
    """Download Excel bytes from SharePoint with browser-like session flow."""
    data, _meta = download_excel_with_meta(url)
    return data


def download_excel_with_meta(url: str) -> tuple[bytes, dict[str, str]]:
    """Download Excel bytes and return response metadata useful for change detection."""
    if not url:
        raise ValueError("EXCEL_URL is empty")
    session = requests.Session()
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
        ),
        "Accept": (
            "text/html,application/xhtml+xml,application/xml;q=0.9,"
            "application/vnd.ms-excel,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*;q=0.8"
        ),
        "Accept-Language": "en-US,en;q=0.9,ar;q=0.8",
        "Cache-Control": "no-cache",
        "Pragma": "no-cache",
    }

    warmup_url = _normalize_sharepoint_download_url(url, cache_bust=True)
    warmup = session.get(warmup_url, headers=headers, allow_redirects=True, timeout=30)
    warmup.raise_for_status()

    requested_url = _normalize_sharepoint_download_url(url, cache_bust=True)
    r = session.get(requested_url, headers=headers, allow_redirects=True, timeout=60)
    r.raise_for_status()

    redirect_urls = [resp.url for resp in r.history] + [r.url]
    final_host = (urlparse(r.url).netloc or "").lower()

    data = r.content or b""
    ctype = (r.headers.get("Content-Type") or "").lower()
    sig16 = _file_signature_hex16(data)
    meta = {
        "etag": (r.headers.get("ETag") or r.headers.get("Etag") or "").strip(),
        "last_modified": (r.headers.get("Last-Modified") or "").strip(),
        "content_length": str(len(data)),
        "content_type": ctype,
    }

    print(f"  Requested URL: {requested_url}")
    print(f"  Final URL: {r.url}")
    print("  Redirect chain:")
    for idx, u in enumerate(redirect_urls, start=1):
        print(f"    {idx}. {u}")
    print(f"  Content-Type: {ctype or 'unknown'}")
    print(f"  Last-Modified: {meta['last_modified'] or 'n/a'}")
    print(f"  ETag: {meta['etag'] or 'n/a'}")
    print(f"  First 16 bytes hex: {sig16}")
    print(f"  File size: {len(data):,} bytes")

    if "login.microsoftonline.com" in final_host:
        raise ValueError("Reached login.microsoftonline.com. Check sharing link and direct download URL.")

    if _is_png_signature(data):
        with open(DEBUG_SHAREPOINT_RESPONSE_PATH, "wb") as f:
            f.write(data)
        raise ValueError("SharePoint returned a preview image, not the Excel file. Use a direct download link.")

    if not _is_excel_signature(data):
        raise ValueError(
            f"Downloaded file is not recognized as Excel payload (Content-Type: {ctype or 'unknown'}; signature: {sig16})"
        )

    return data, meta


def download_text(url: str) -> str:
    bust = _add_or_replace_query_param(url, "_cb", str(int(time.time() * 1000))) if url else url
    r = requests.get(
        bust or url,
        timeout=30,
        headers={"Cache-Control": "no-cache", "Pragma": "no-cache"},
    )
    r.raise_for_status()
    return r.text.strip()


def get_source_name() -> str:
    if SOURCE_NAME_URL:
        try:
            name = download_text(SOURCE_NAME_URL)
            if name:
                return name
        except Exception:
            pass
    return SOURCE_NAME_FALLBACK or "latest.xlsx"


def infer_pages_base_url() -> str:
    return "https://khalidsaif912.github.io/roster-site"


MONTH_NAME_TO_NUM = {
    "january": 1,
    "jan": 1,
    "february": 2,
    "feb": 2,
    "march": 3,
    "mar": 3,
    "april": 4,
    "apr": 4,
    "may": 5,
    "june": 6,
    "jun": 6,
    "july": 7,
    "jul": 7,
    "august": 8,
    "aug": 8,
    "september": 9,
    "sep": 9,
    "sept": 9,
    "october": 10,
    "oct": 10,
    "november": 11,
    "nov": 11,
    "december": 12,
    "dec": 12,
}


_ROSTER_MONTH_HINT = re.compile(
    r"\b(january|jan|february|feb|march|mar|april|apr|may|june|jun|july|jul|august|aug|september|sep|sept|october|oct|november|nov|december|dec)\b|20\d{2}",
    re.IGNORECASE,
)


def looks_like_roster_month_filename(name: str) -> bool:
    if not name:
        return False
    return bool(_ROSTER_MONTH_HINT.search(name))


def month_key_from_filename(name: str) -> str | None:
    if not name:
        return None
    n = name.lower()
    n = re.sub(r"[\._\-]+", " ", n)
    n = re.sub(r"\s+", " ", n).strip()
    # YYYY-MM or YYYY_MM in filename
    m = re.search(r"\b(20\d{2})[-_ ](0[1-9]|1[0-2])\b", n)
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}"
    # Month name + year (space optional, e.g. "July 2026" or "July2026")
    m = re.search(
        r"\b(january|jan|february|feb|march|mar|april|apr|may|june|jun|july|jul|august|aug|september|sep|sept|october|oct|november|nov|december|dec)[\s_-]*(20\d{2})\b",
        n,
    )
    if not m:
        return None
    mon_name, year_s = m.group(1), m.group(2)
    mon = MONTH_NAME_TO_NUM.get(mon_name)
    if not mon:
        return None
    return f"{int(year_s):04d}-{mon:02d}"


def add_months(year: int, month: int, delta: int) -> tuple[int, int]:
    y = year
    m = month + delta
    while m <= 0:
        y -= 1
        m += 12
    while m > 12:
        y += 1
        m -= 12
    return y, m


def cache_paths(month_key: str) -> tuple[str, str]:
    os.makedirs(ROSTERS_DIR, exist_ok=True)
    return (
        os.path.join(ROSTERS_DIR, f"{month_key}.xlsx"),
        os.path.join(ROSTERS_DIR, f"{month_key}.meta.json"),
    )


def write_bytes(path: str, data: bytes):
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    with open(path, "wb") as f:
        f.write(data)


def read_json(path: str) -> dict | None:
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None


def write_json(path: str, obj: dict):
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False, indent=2)


def try_load_cached_workbook(month_key: str):
    xlsx_path, _ = cache_paths(month_key)
    if not os.path.exists(xlsx_path):
        return None
    try:
        with open(xlsx_path, "rb") as f:
            return load_workbook(BytesIO(f.read()), data_only=True)
    except Exception:
        return None


def cached_source_name(month_key: str) -> str:
    _, meta_path = cache_paths(month_key)
    meta = read_json(meta_path) or {}
    return (meta.get("original_filename") or meta.get("source_name") or "").strip()
