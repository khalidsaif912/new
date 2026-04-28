#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import load_workbook


DAYS = {"SUN", "MON", "TUE", "WED", "THU", "FRI", "SAT"}
SKIP_SHEETS = {"full staffs as per jd"}


def norm(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def is_emp_id(v: str) -> bool:
    v = norm(v)
    return bool(re.fullmatch(r"\d{3,}", v))


def looks_name(v: str) -> bool:
    v = norm(v)
    if not v or is_emp_id(v):
        return False
    return any(ch.isalpha() for ch in v)


def find_day_row(ws) -> int:
    max_scan = min(80, ws.max_row)
    for r in range(1, max_scan + 1):
        vals = [norm(ws.cell(row=r, column=c).value).upper() for c in range(1, ws.max_column + 1)]
        tokens = 0
        for x in vals:
            if any(d in x for d in DAYS):
                tokens += 1
        if tokens >= 3:
            return r
    raise ValueError("Could not detect day header row (SUN/MON/...)")


def day_cols(ws, day_row: int) -> List[Tuple[int, int]]:
    cols = []
    idx = 1
    for c in range(1, ws.max_column + 1):
        v = norm(ws.cell(row=day_row, column=c).value).upper()
        if any(d in v for d in DAYS):
            cols.append((idx, c))
            idx += 1
    if not cols:
        raise ValueError("No day columns found")
    return cols


def parse_file(path: Path) -> Dict[str, Dict]:
    wb = load_workbook(path, data_only=True)
    out: Dict[str, Dict] = {}
    parsed_any_sheet = False

    for sn in wb.sheetnames:
        if norm(sn).lower() in SKIP_SHEETS:
            continue
        ws = wb[sn]
        try:
            drow = find_day_row(ws)
            dcols = day_cols(ws, drow)
        except Exception:
            continue
        parsed_any_sheet = True

        for r in range(drow + 1, ws.max_row + 1):
            row_vals = [norm(ws.cell(row=r, column=c).value) for c in range(1, ws.max_column + 1)]
            emp_id = ""
            name = ""
            for i, v in enumerate(row_vals):
                if not emp_id and is_emp_id(v):
                    emp_id = v
                    if i + 1 < len(row_vals) and looks_name(row_vals[i + 1]):
                        name = row_vals[i + 1]
                    elif i - 1 >= 0 and looks_name(row_vals[i - 1]):
                        name = row_vals[i - 1]
                    break
            # Some local roster templates do not include numeric employee IDs.
            # In that case, fall back to the employee name as a stable key
            # so auto-diff still works for shift changes.
            if not name:
                for v in row_vals:
                    if looks_name(v):
                        name = v
                        break

            shift_map: Dict[str, str] = {}
            for day_num, col in dcols:
                code = norm(ws.cell(row=r, column=col).value).upper()
                if code:
                    shift_map[str(day_num)] = code

            # Skip totals/summary rows that don't contain shift codes.
            has_alpha_shift = any(any(ch.isalpha() for ch in code) for code in shift_map.values())
            if not has_alpha_shift:
                continue

            key = emp_id or f"name:{name}"
            if key == "name:":
                continue
            out.setdefault(key, {"name": name, "shifts": {}})
            if name and not out[key]["name"]:
                out[key]["name"] = name
            out[key]["shifts"].update(shift_map)

    if not parsed_any_sheet:
        raise ValueError("Could not detect day header row in any sheet")
    return out


def build_diff(old_path: Path, new_path: Path) -> List[Dict]:
    old_data = parse_file(old_path)
    new_data = parse_file(new_path)
    changes: List[Dict] = []
    ids = sorted(set(old_data.keys()) | set(new_data.keys()))
    for emp_id in ids:
        a = old_data.get(emp_id, {"name": "", "shifts": {}})
        b = new_data.get(emp_id, {"name": "", "shifts": {}})
        days = sorted(set(a["shifts"].keys()) | set(b["shifts"].keys()), key=lambda x: int(x))
        for day in days:
            v1 = a["shifts"].get(day, "")
            v2 = b["shifts"].get(day, "")
            if v1 != v2:
                changes.append(
                    {
                        "emp_id": emp_id,
                        "name": b.get("name") or a.get("name") or "",
                        "day": int(day),
                        "v1": v1,
                        "v2": v2,
                    }
                )
    return changes


def main() -> None:
    p = argparse.ArgumentParser(description="Build roster diff JSON from old/new xlsx")
    p.add_argument("--old", required=True)
    p.add_argument("--new", required=True)
    p.add_argument("--kind", choices=["export", "import"], required=True)
    p.add_argument("--month", required=True, help="YYYY-MM")
    p.add_argument("--out-dir", default="docs/roster-diff/data")
    args = p.parse_args()

    old_path = Path(args.old)
    new_path = Path(args.new)
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    if not re.fullmatch(r"\d{4}-\d{2}", args.month):
        raise SystemExit("month must be YYYY-MM")
    if not old_path.exists() or not new_path.exists():
        raise SystemExit("old/new file missing")

    changes = build_diff(old_path, new_path)
    payload = {
        "kind": args.kind,
        "month": args.month,
        "generated_at": datetime.utcnow().isoformat() + "Z",
        "old_file": old_path.name,
        "new_file": new_path.name,
        "total_changes": len(changes),
        "changes": changes,
    }

    month_file = out_dir / f"{args.kind}-{args.month}.json"
    latest_file = out_dir / f"{args.kind}-latest.json"
    month_file.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    latest_file.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"OK diff: {latest_file}")


if __name__ == "__main__":
    main()
