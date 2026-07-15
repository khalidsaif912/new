#!/usr/bin/env python3
"""Compare oldest vs current export/import rosters to find people who left."""
from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
TRANS = json.loads((ROOT / "docs" / "name_translations.json").read_text(encoding="utf-8"))
NAMES = TRANS.get("names") or {}

SKIP_SHEETS = {"setting", "setting ", "master", "full staffs as per jd"}
NAME_RE = re.compile(r"^\s*(.+?)\s*[-–]\s*(\d{3,})\s*$")

JD_MAP = {
    "SUPV": "Supervisors",
    "FLTI": "Flight Dispatch (Import)",
    "FLTE": "Flight Dispatch (Export)",
    "DOC": "Documentation",
    "DOCS": "Documentation",
    "ICHK": "Import Checkers",
    "IOPS": "Import Operators",
    "RC": "Release Control",
    "CHK": "Import Checkers",
    "OPS": "Import Operators",
}


def ar_for(en: str) -> str:
    key = re.sub(r"\s+", " ", en).strip().upper()
    return NAMES.get(key) or ""


def extract_export(path: Path) -> dict:
    wb = load_workbook(path, data_only=True, read_only=True)
    out: dict = {}
    for sheet in wb.sheetnames:
        if sheet.strip().lower() in SKIP_SHEETS:
            continue
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            for cell in row[:6]:
                if not isinstance(cell, str):
                    continue
                m = NAME_RE.match(cell)
                if not m:
                    continue
                name, eid = m.group(1).strip(), m.group(2)
                if eid not in out:
                    out[eid] = {
                        "id": eid,
                        "name": name,
                        "department": sheet.strip(),
                        "kind": "export",
                    }
    wb.close()
    return out


def extract_import(path: Path) -> dict:
    wb = load_workbook(path, data_only=True, read_only=True)
    out: dict = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        header_i = None
        for i, row in enumerate(rows[:10]):
            vals = [str(c).strip().lower() if c is not None else "" for c in row[:5]]
            if "employee name" in vals:
                header_i = i
                break
        if header_i is None:
            continue
        for row in rows[header_i + 1 :]:
            if not row or len(row) < 3:
                continue
            jd = str(row[0] or "").strip()
            name = str(row[1] or "").strip()
            sn = str(row[2] or "").strip()
            if not name or not re.fullmatch(r"\d{3,}", sn):
                continue
            name_disp = name.title() if name.isupper() else name
            dept = JD_MAP.get(jd.upper(), jd or "Import")
            if sn not in out:
                out[sn] = {
                    "id": sn,
                    "name": name_disp,
                    "department": dept,
                    "kind": "import",
                    "jd": jd,
                }
    wb.close()
    return out


def main() -> None:
    old_exp = extract_export(ROOT / "rosters" / "2026-02.xlsx")
    new_exp = extract_export(ROOT / "rosters" / "2026-07.xlsx")
    old_imp = extract_import(ROOT / "import-rosters" / "2026-03.xlsx")
    new_imp = extract_import(ROOT / "import-rosters" / "2026-07.xlsx")

    left_exp = []
    for eid, e in sorted(old_exp.items(), key=lambda x: x[1]["name"].lower()):
        if eid not in new_exp:
            status = "moved_to_import" if eid in new_imp else "left"
            left_exp.append({**e, "status": status, "nameAr": ar_for(e["name"])})

    left_imp = []
    for eid, e in sorted(old_imp.items(), key=lambda x: x[1]["name"].lower()):
        if eid not in new_imp:
            status = "moved_to_export" if eid in new_exp else "left"
            left_imp.append({**e, "status": status, "nameAr": ar_for(e["name"])})

    print(f"EXPORT old={len(old_exp)} new={len(new_exp)}")
    print(f"IMPORT old={len(old_imp)} new={len(new_imp)}")
    print("\n=== EXPORT not in current export ===")
    for e in left_exp:
        tag = " [MOVED IMPORT]" if e["status"] != "left" else ""
        print(f"{e['id']:>8}  {e['name']:<35}  {e['department']:<20}  {e['nameAr']}{tag}")
    print(
        "total",
        len(left_exp),
        "truly left",
        sum(1 for e in left_exp if e["status"] == "left"),
    )

    print("\n=== IMPORT not in current import ===")
    for e in left_imp:
        tag = " [MOVED EXPORT]" if e["status"] != "left" else ""
        print(f"{e['id']:>8}  {e['name']:<35}  {e['department']:<28}  {e['nameAr']}{tag}")
    print(
        "total",
        len(left_imp),
        "truly left",
        sum(1 for e in left_imp if e["status"] == "left"),
    )

    payload = {
        "export": {
            "old_file": "rosters/2026-02.xlsx",
            "new_file": "rosters/2026-07.xlsx",
            "left": [e for e in left_exp if e["status"] == "left"],
            "moved": [e for e in left_exp if e["status"] != "left"],
        },
        "import": {
            "old_file": "import-rosters/2026-03.xlsx",
            "new_file": "import-rosters/2026-07.xlsx",
            "left": [e for e in left_imp if e["status"] == "left"],
            "moved": [e for e in left_imp if e["status"] != "left"],
        },
    }
    out = ROOT / "docs" / "_alumni_compare.json"
    out.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print("wrote", out)


if __name__ == "__main__":
    main()
