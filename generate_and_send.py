import os
import re
import json
import sys
import calendar
import argparse
from pathlib import Path
from html import escape as html_escape

_SCRIPTS_DIR = Path(__file__).resolve().parent / "scripts"
if str(_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS_DIR))
from roster_cta_snippets import (  # noqa: E402
    CHIP_AFTERNOON_HTML,
    CHIP_ALL_HTML,
    CHIP_DIFF_HTML,
    CHIP_FLIGHT_HTML,
    CHIP_ICON_CSS,
    CHIP_MORNING_HTML,
    CHIP_NIGHT_HTML,
    CHIP_SCHEDULE_HTML,
    CHIP_TRAINING_HTML,
    CHIP_WAVE_HTML,
    LANG_TOGGLE_CSS,
    LANG_TOGGLE_HTML,
    APPLY_LANG_LANG_BTN_NEW,
    SITE_APPS_MODAL_HTML,
    SITE_SHARE_MODAL_HTML,
    SHIFT_COPY_BUTTON_HTML,
    SHIFT_COPY_MODAL_HTML,
    SHIFT_COPY_CSS,
    export_cta_html,
    IOS_PERF_VER,
    LOAD_LOCAL_ENHANCEMENTS_EXPORT,
    PERF_RENDER_CSS,
)
from datetime import datetime
from io import BytesIO

from openpyxl import load_workbook
from roster_app.cache_io import (
    add_months,
    cache_paths,
    cached_source_name,
    download_excel,
    get_source_name,
    infer_pages_base_url,
    looks_like_roster_month_filename,
    month_key_from_filename,
    try_load_cached_workbook,
    write_bytes,
    write_json,
)
from roster_app.email_service import send_email
from roster_app.settings import (
    AUTO_OPEN_ACTIVE_SHIFT_IN_FULL,
    DAYS,
    DEPARTMENTS,
    EXCEL_URL,
    GROUP_ORDER,
    PAGES_BASE_URL,
    SHIFT_MAP,
    TZ,
)
from roster_app.text_utils import (
    append_range_suffix,
    clean,
    current_shift_key,
    looks_like_employee_name,
    looks_like_shift_code,
    looks_like_time,
    map_shift,
    norm,
    range_suffix_for_day,
    to_western_digits,
)
from roster_app import name_i18n

# PWA: manifest + iOS meta; paths resolve for /roster-site/ and GitHub Pages /user/docs/
ROSTER_PWA_HEAD_SNIPPET = """
  <meta name="theme-color" content="#f4354b">
  <meta name="apple-mobile-web-app-capable" content="yes">
  <meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
  <meta name="mobile-web-app-capable" content="yes">
  <script>
  (function () {
    function siteRoot() {
      if (location.protocol === 'file:') return '';
      var path = location.pathname || '/';
      if (path.indexOf('/roster-site/') !== -1) return '/roster-site';
      if (location.hostname && location.hostname.endsWith('github.io')) {
        var segs = path.split('/').filter(Boolean);
        if (segs.length >= 2 && segs[1] === 'docs') return '/' + segs[0] + '/docs';
        return segs.length ? '/' + segs[0] : '';
      }
      return '';
    }
    var p = siteRoot();
    var base = location.origin + p + (p && p.charAt(p.length - 1) !== '/' ? '/' : '');
    if (!p) base = location.origin + '/';
    var pv = '14';
    var imp = (location.pathname || '').indexOf('/import/') !== -1;
    var man = base + (imp ? 'import/manifest.json' : 'manifest.json') + '?v=' + pv;
    var mlinks = document.querySelectorAll('link[rel="manifest"]');
    var link = mlinks.length ? mlinks[0] : null;
    if (!link) {
      link = document.createElement('link');
      link.rel = 'manifest';
      document.head.appendChild(link);
    }
    link.href = man;
    for (var i = 1; i < mlinks.length; i++) mlinks[i].remove();
    var touch = document.querySelector('link[rel="apple-touch-icon"][data-pwa-touch="1"]');
    if (!touch) {
      touch = document.createElement('link');
      touch.rel = 'apple-touch-icon';
      touch.setAttribute('data-pwa-touch', '1');
      document.head.appendChild(touch);
    }
    touch.href = base + 'assets/icons/icon-192.png';
    try {{
      var iOS = /iP(hone|ad|od)/i.test(navigator.userAgent) ||
        (navigator.platform === 'MacIntel' && navigator.maxTouchPoints > 1);
      if (iOS && 'serviceWorker' in navigator) {{
        navigator.serviceWorker.getRegistrations().then(function(regs) {{
          regs.forEach(function(r) {{ r.unregister(); }});
        }});
      }}
    }} catch (swIosErr) {{}}
    try {{
      var bn = localStorage.getItem('roster_banner_choice');
      if (bn && /^banner\\d+\\.jpg$/i.test(bn)) {{
        var bUrl = base + 'assets/banners/' + bn;
        if (!document.getElementById('banner-early-style')) {{
          var bes = document.createElement('style');
          bes.id = 'banner-early-style';
          bes.textContent =
            'html.roster-banner-early .header,html.roster-banner-early .topbar{{background-image:url("' + bUrl.replace(/"/g, '') + '")!important;background-size:cover!important;background-position:62% center!important;background-repeat:no-repeat!important}}' +
            'html.roster-banner-early .header::before,html.roster-banner-early .header::after{{opacity:0!important}}';
          document.head.appendChild(bes);
        }}
        if (!document.querySelector('link[data-banner-preload="1"]')) {{
          var bp = document.createElement('link');
          bp.rel = 'preload';
          bp.as = 'image';
          bp.href = bUrl;
          bp.setAttribute('data-banner-preload', '1');
          document.head.appendChild(bp);
        }}
        document.documentElement.classList.add('roster-banner-early');
      }}
    }} catch (bannerEarlyErr) {{}}
  }})();
  </script>
"""


def format_site_last_updated(dt: datetime) -> str:
    """Readable Muscat publish time for site footer (not roster calendar day)."""
    try:
        day = dt.strftime("%-d")
    except ValueError:
        day = dt.strftime("%d").lstrip("0") or dt.strftime("%d")
    return f"{day} {dt.strftime('%B %Y')} / {dt.strftime('%H:%M')}"


def write_site_last_updated_json(dt: datetime) -> None:
    display_en = format_site_last_updated(dt)
    write_json(
        "docs/site-last-updated.json",
        {
            "updated_at": dt.isoformat(),
            "display_en": display_en,
            "display_ar": display_en,
        },
    )


# =========================
# Detect rows/cols (Days row + Date numbers row)
# =========================
def _row_values(ws, r: int):
    return [norm(ws.cell(row=r, column=c).value) for c in range(1, ws.max_column + 1)]

def _count_day_tokens(vals) -> int:
    ups = [v.upper() for v in vals if v]
    count = 0
    for d in DAYS:
        if any(d in x for x in ups):
            count += 1
    return count

def _is_date_number(v: str) -> bool:
    v = norm(v)
    if not v:
        return False
    if re.match(r"^\d{1,2}(\.0)?$", v):
        n = int(float(v))
        return 1 <= n <= 31
    return False

def find_days_and_dates_rows(ws, scan_rows: int = 80):
    """
    يبحث عن صف فيه SUN..SAT بكثرة ثم صف تحته فيه أرقام 1..31
    """
    max_r = min(ws.max_row, scan_rows)
    days_row = None

    for r in range(1, max_r + 1):
        vals = _row_values(ws, r)
        if _count_day_tokens(vals) >= 3:
            days_row = r
            break

    if not days_row:
        return None, None

    date_row = None
    for r in range(days_row + 1, min(days_row + 4, ws.max_row) + 1):
        vals = _row_values(ws, r)
        nums = sum(1 for v in vals if _is_date_number(v))
        if nums >= 5:
            date_row = r
            break

    return days_row, date_row

def find_day_col(ws, days_row: int, date_row: int, today_dow: int, today_day: int):
    """
    يثبت العمود الصحيح باستخدام اليوم + رقم التاريخ
    """
    if not days_row or not date_row:
        return None

    day_key = DAYS[today_dow]
    # Prefer (day + date) match
    for c in range(1, ws.max_column + 1):
        top = norm(ws.cell(row=days_row, column=c).value).upper()
        bot = norm(ws.cell(row=date_row, column=c).value)
        if day_key in top and _is_date_number(bot) and int(float(bot)) == today_day:
            return c

    # Fallback: date-only
    for c in range(1, ws.max_column + 1):
        bot = norm(ws.cell(row=date_row, column=c).value)
        if _is_date_number(bot) and int(float(bot)) == today_day:
            return c

    return None


def get_daynum_to_col(ws, date_row: int):
    m = {}
    for c in range(1, ws.max_column + 1):
        v = norm(ws.cell(row=date_row, column=c).value)
        if _is_date_number(v):
            m[int(float(v))] = c
    return m

def find_employee_col(ws, start_row: int):
    for c in range(1, min(ws.max_column, 15) + 1):
        found = 0
        for r in range(start_row, min(start_row + 20, ws.max_row) + 1):
            v = norm(ws.cell(row=r, column=c).value)
            if looks_like_employee_name(v):
                found += 1
        if found >= 3:
            return c
    return None

def _apply_shift_range_label(label: str, grp: str, day: int, daynum_to_raw: dict, raw: str) -> str:
    """Apply (FROM x TO y) for leave blocks (suffix only) and regular multi-day shifts (code + suffix)."""
    up = norm(raw).upper()
    if grp == "Annual Leave" and (up == "AL" or "ANNUAL LEAVE" in up or up == "LV"):
        suf = range_suffix_for_day(day, daynum_to_raw, raw)
        return suf or label
    if grp == "Sick Leave" and (up == "SL" or "SICK LEAVE" in up):
        suf = range_suffix_for_day(day, daynum_to_raw, raw)
        return suf or label
    if grp == "Training" and (up == "TR" or "TRAINING" in up):
        suf = range_suffix_for_day(day, daynum_to_raw, raw)
        return suf or label
    return label


# =========================
# Department card colors
# =========================
DEPT_COLORS = [
    {"name": "blue",   "base": "#2563eb", "light": "#2563eb15", "border": "#2563eb18", "grad_from": "#2563eb", "grad_to": "#2563ebcc"},
    {"name": "cyan",   "base": "#0891b2", "light": "#0891b215", "border": "#0891b218", "grad_from": "#0891b2", "grad_to": "#0891b2cc"},
    {"name": "green",  "base": "#059669", "light": "#05966915", "border": "#05966918", "grad_from": "#059669", "grad_to": "#059669cc"},
    {"name": "red",    "base": "#dc2626", "light": "#dc262615", "border": "#dc262618", "grad_from": "#dc2626", "grad_to": "#dc2626cc"},
    {"name": "purple", "base": "#7c3aed", "light": "#7c3aed15", "border": "#7c3aed18", "grad_from": "#7c3aed", "grad_to": "#7c3aedcc"},
    {"name": "orange", "base": "#ea580c", "light": "#ea580c15", "border": "#ea580c18", "grad_from": "#ea580c", "grad_to": "#ea580ccc"},
]

# قسم Unassigned يأخذ لون برتقالي/رمادي
UNASSIGNED_COLOR = {"name": "gray", "base": "#6b7280", "light": "#6b728015", "border": "#6b728018", "grad_from": "#6b7280", "grad_to": "#6b7280cc"}

# =========================
# Shift group colors (Morning/Afternoon/Night/etc.)
# =========================
SHIFT_COLORS = {
    "Morning": {
        "border": "#f59e0b44",
        "bg": "#fef3c7",
        "summary_bg": "#fef3c7",
        "summary_border": "#f59e0b33",
        "label_color": "#92400e",
        "count_bg": "#f59e0b22",
        "count_color": "#92400e",
        "status_color": "#92400e",
        "icon": "☀️",
    },
    "Afternoon": {
        "border": "#f9731644",
        "bg": "#ffedd5",
        "summary_bg": "#ffedd5",
        "summary_border": "#f9731633",
        "label_color": "#9a3412",
        "count_bg": "#f9731622",
        "count_color": "#9a3412",
        "status_color": "#9a3412",
        "icon": "🌤️",
    },
    "Night": {
        "border": "#8b5cf644",
        "bg": "#ede9fe",
        "summary_bg": "#ede9fe",
        "summary_border": "#8b5cf633",
        "label_color": "#5b21b6",
        "count_bg": "#8b5cf622",
        "count_color": "#5b21b6",
        "status_color": "#5b21b6",
        "icon": "🌙",
    },
    "Off Day": {
        "border": "#6366f144",
        "bg": "#e0e7ff",
        "summary_bg": "#e0e7ff",
        "summary_border": "#6366f133",
        "label_color": "#3730a3",
        "count_bg": "#6366f122",
        "count_color": "#3730a3",
        "status_color": "#3730a3",
        "icon": "🛋️",
    },
    "Annual Leave": {
        "border": "#10b98144",
        "bg": "#d1fae5",
        "summary_bg": "#d1fae5",
        "summary_border": "#10b98133",
        "label_color": "#065f46",
        "count_bg": "#10b98122",
        "count_color": "#065f46",
        "status_color": "#065f46",
        "icon": "✈️",
    },
    "Training": {
        "border": "#0ea5e944",
        "bg": "#e0f2fe",
        "summary_bg": "#e0f2fe",
        "summary_border": "#0ea5e933",
        "label_color": "#075985",
        "count_bg": "#0ea5e922",
        "count_color": "#075985",
        "status_color": "#075985",
        "icon": "📚",
    },
    "Standby": {
        "border": "#9e9e9e44",
        "bg": "#f0f0f0",
        "summary_bg": "#f0f0f0",
        "summary_border": "#9e9e9e33",
        "label_color": "#555555",
        "count_bg": "#cccccc22",
        "count_color": "#555555",
        "status_color": "#555555",
        "icon": "🧍"
    }, 
    "Sick Leave": {
    "border": "#ef444444",
    "bg": "#fee2e2",
    "summary_bg": "#fee2e2",
    "summary_border": "#ef444433",
    "label_color": "#991b1b",
    "count_bg": "#ef444422",
    "count_color": "#991b1b",
    "status_color": "#991b1b",
    "icon": "🤒",
   },
    "Other": {
        "border": "#94a3b844",
        "bg": "#f1f5f9",
        "summary_bg": "#f1f5f9",
        "summary_border": "#94a3b833",
        "label_color": "#475569",
        "count_bg": "#94a3b822",
        "count_color": "#475569",
        "status_color": "#475569",
        "icon": "❓",
    },
}


# Next 5 shifts tooltip: long-press on .empRow (touch + mouse); tap → schedule (injected into page_shell_html).
EMPLOYEE_NEXT_SHIFT_PREVIEW_JS = """
// Employee row: tap → schedule; long-press → next 5 shifts preview
(function initEmployeeNextShiftPreview() {
  var tooltip = document.createElement('div');
  tooltip.className = 'nextShiftTooltip';
  tooltip.innerHTML = '<div class="nextShiftHead"><div class="nextShiftHeadText"><div id="nextShiftEmp" class="nextShiftEmp">-</div><div class="nextShiftTitle">Upcoming 5 shifts</div></div><button type="button" class="nextShiftClose" id="nextShiftClose" aria-label="Close">&times;</button></div><div id="nextShiftBody" class="nextShiftBody"></div>';
  document.body.appendChild(tooltip);
  var tooltipBody = tooltip.querySelector('#nextShiftBody');
  var tooltipEmp = tooltip.querySelector('#nextShiftEmp');
  var tooltipClose = tooltip.querySelector('#nextShiftClose');

  var scheduleCache = {};
  var tooltipPinned = false;
  var activeEl = null;
  var hideTimer = null;
  var longPressTimer = null;
  var suppressClickFor = null;
  var longPressRow = null;
  var longPressMoved = false;
  var LONG_PRESS_MS = 550;

  function empNameFromRow(rowEl) {
    if (!rowEl) return '';
    var raw = rowEl.getAttribute('data-emp-name');
    if (raw) return raw.trim();
    var nameEl = rowEl.querySelector('.empName');
    return nameEl ? String(nameEl.textContent || '').trim() : '';
  }

  function getReferenceIsoDate() {
    var pathMatch = (location.pathname || '').match(/\\/date\\/(\\d{4}-\\d{2}-\\d{2})\\//);
    if (pathMatch) return pathMatch[1];
    var picker = document.getElementById('datePicker');
    if (picker && picker.value) return picker.value;
    var now = new Date();
    var muscat = new Date(now.getTime() + (4 * 60 * 60 * 1000) + (now.getTimezoneOffset() * 60 * 1000));
    return muscat.getFullYear() + '-' +
      String(muscat.getMonth() + 1).padStart(2, '0') + '-' +
      String(muscat.getDate()).padStart(2, '0');
  }

  function moveTooltip(ev) {
    if (!ev) return;
    var x = ev.clientX + 14;
    var y = ev.clientY + 14;
    var maxX = window.innerWidth - tooltip.offsetWidth - 10;
    var maxY = window.innerHeight - tooltip.offsetHeight - 10;
    tooltip.style.left = Math.max(8, Math.min(x, maxX)) + 'px';
    tooltip.style.top = Math.max(8, Math.min(y, maxY)) + 'px';
  }
  function moveTooltipToElement(el) {
    if (!el) return;
    var rect = el.getBoundingClientRect();
    var x = rect.left + rect.width / 2 + 12;
    var y = rect.top + rect.height + 12;
    var maxX = window.innerWidth - tooltip.offsetWidth - 10;
    var maxY = window.innerHeight - tooltip.offsetHeight - 10;
    tooltip.style.left = Math.max(8, Math.min(x, maxX)) + 'px';
    tooltip.style.top = Math.max(8, Math.min(y, maxY)) + 'px';
  }

  function cancelHideTooltip() {
    if (hideTimer) {
      clearTimeout(hideTimer);
      hideTimer = null;
    }
  }

  function hideTooltipNow() {
    cancelHideTooltip();
    tooltipPinned = false;
    tooltip.classList.remove('show');
    activeEl = null;
  }

  function hideTooltipSoon() {
    if (tooltipPinned) return;
    cancelHideTooltip();
    hideTimer = setTimeout(function() {
      tooltip.classList.remove('show');
      activeEl = null;
      hideTimer = null;
    }, 120);
  }

  function parseEmployeeId(nameText) {
    var match = String(nameText || '').match(/-\\s*(\\d{3,})/);
    return match ? match[1] : '';
  }

  function flattenFutureShifts(data, fromIso) {
    var out = [];
    if (!data) return out;
    if (!data.schedules) return out;
    Object.keys(data.schedules).forEach(function(monthKey) {
      var mk = String(monthKey).match(/^(\\d{4})-(\\d{2})$/);
      if (!mk) return;
      var y = mk[1], mo = mk[2];
      var rows = data.schedules[monthKey] || [];
      rows.forEach(function(r) {
        if (!r) return;
        var iso = String(r.date || '').trim();
        if (!iso && r.day != null && r.day !== '') {
          iso = y + '-' + mo + '-' + String(r.day).padStart(2, '0');
        }
        if (!iso || iso < fromIso) return;
        out.push({ date: iso, shift_code: String(r.shift_code || r.code || '').trim() });
      });
    });
    out.sort(function(a, b) {
      return String(a.date).localeCompare(String(b.date));
    });
    return out.slice(0, 5);
  }

  function formatShortDateParts(isoDate) {
    var match = String(isoDate || '').match(/^(\\d{4})-(\\d{2})-(\\d{2})$/);
    if (!match) return { dayName: '--', dateLabel: String(isoDate || '-') };
    var d = new Date(Number(match[1]), Number(match[2]) - 1, Number(match[3]));
    if (isNaN(d.getTime())) return { dayName: '--', dateLabel: String(isoDate || '-') };
    var dayName = d.toLocaleDateString('en-GB', { weekday: 'short' }).toUpperCase();
    var dateLabel = d.toLocaleDateString('en-GB', { day: '2-digit', month: 'short' });
    return { dayName: dayName, dateLabel: dateLabel };
  }

  function renderTooltipRows(rows) {
    if (!rows.length) {
      tooltipBody.innerHTML = '<div class="nextShiftEmpty">No upcoming shifts found.</div>';
      return;
    }
    tooltipBody.innerHTML = rows.map(function(r) {
      var parts = formatShortDateParts(String(r.date || ''));
      var shiftCode = String(r.shift_code || '').trim() || '-';
      return '<div class="nextShiftItem"><div class="nextShiftDay">' + parts.dayName + '</div><div class="nextShiftDate">' + parts.dateLabel + '</div><div class="nextShiftCode">' + shiftCode + '</div></div>';
    }).join('');
  }

  function getScheduleRows(empId, fromIso) {
    if (!empId) return Promise.resolve([]);
    if (scheduleCache[empId]) {
      return Promise.resolve(flattenFutureShifts(scheduleCache[empId], fromIso));
    }
    var url = getSiteRootUrl() + '/schedules/' + encodeURIComponent(empId) + '.json';
    return fetch(url).then(function(res) {
      if (!res.ok) throw new Error('Schedule not found');
      return res.json();
    }).then(function(json) {
      scheduleCache[empId] = json;
      return flattenFutureShifts(json, fromIso);
    }).catch(function() {
      return [];
    });
  }

  function showPreviewForRow(rowEl, ev) {
    if (!rowEl) return;
    cancelHideTooltip();
    tooltipPinned = true;
    activeEl = rowEl;
    var label = empNameFromRow(rowEl);
    if (tooltipEmp) tooltipEmp.textContent = label || '-';
    tooltipBody.innerHTML = '<div class="nextShiftEmpty">Loading...</div>';
    tooltip.classList.add('show');
    if (ev && typeof ev.clientX === 'number') moveTooltip(ev);
    else moveTooltipToElement(rowEl);

    var empId = parseEmployeeId(label);
    var fromIso = getReferenceIsoDate();
    getScheduleRows(empId, fromIso).then(function(rows) {
      if (activeEl !== rowEl) return;
      renderTooltipRows(rows);
      if (ev && typeof ev.clientX === 'number') moveTooltip(ev);
      else moveTooltipToElement(rowEl);
    });
  }

  function bindEmployeeRow(rowEl) {
    if (!rowEl || rowEl.dataset.nextShiftBound === '1') return;
    rowEl.dataset.nextShiftBound = '1';

    rowEl.addEventListener('pointerdown', function(ev) {
      if (ev.button !== 0 && ev.button !== undefined) return;
      longPressRow = rowEl;
      longPressMoved = false;
      if (longPressTimer) clearTimeout(longPressTimer);
      longPressTimer = setTimeout(function() {
        longPressTimer = null;
        if (longPressRow !== rowEl || longPressMoved) return;
        suppressClickFor = rowEl;
        showPreviewForRow(rowEl, null);
      }, LONG_PRESS_MS);
    });
    rowEl.addEventListener('pointermove', function(ev) {
      if (longPressRow !== rowEl) return;
      if (typeof ev.movementX === 'number' && (Math.abs(ev.movementX) > 8 || Math.abs(ev.movementY) > 8)) {
        longPressMoved = true;
      }
    });
    rowEl.addEventListener('pointerup', function() {
      if (longPressTimer) clearTimeout(longPressTimer);
      longPressTimer = null;
      longPressRow = null;
    });
    rowEl.addEventListener('pointercancel', function() {
      if (longPressTimer) clearTimeout(longPressTimer);
      longPressTimer = null;
      longPressRow = null;
    });
    rowEl.addEventListener('contextmenu', function(ev) {
      if (suppressClickFor === rowEl) ev.preventDefault();
    });

    rowEl.addEventListener('click', function(ev) {
      if (suppressClickFor === rowEl) {
        ev.preventDefault();
        ev.stopPropagation();
        suppressClickFor = null;
        return;
      }
      ev.preventDefault();
      goToEmployeeSchedule(empNameFromRow(rowEl));
    }, true);

    rowEl.addEventListener('keydown', function(ev) {
      if (ev.key === 'Enter' || ev.key === ' ') {
        ev.preventDefault();
        goToEmployeeSchedule(empNameFromRow(rowEl));
      }
    });
  }

  if (tooltipClose) {
    tooltipClose.addEventListener('click', function(ev) {
      ev.preventDefault();
      ev.stopPropagation();
      hideTooltipNow();
    });
  }

  function isTooltipOpen() {
    return tooltip.classList.contains('show');
  }

  function dismissUnlessTooltipTarget(ev) {
    if (!isTooltipOpen()) return;
    var t = ev && ev.target;
    if (t && typeof t.closest === 'function' && t.closest('.nextShiftTooltip')) return;
    hideTooltipNow();
  }

  document.addEventListener('pointerdown', dismissUnlessTooltipTarget, true);
  document.addEventListener('click', function(ev) {
    if (suppressClickFor) return;
    dismissUnlessTooltipTarget(ev);
  }, true);

  function dismissOnScroll() {
    if (isTooltipOpen()) hideTooltipNow();
  }
  window.addEventListener('scroll', dismissOnScroll, true);
  window.addEventListener('wheel', dismissOnScroll, { passive: true, capture: true });
  window.addEventListener('touchmove', function(ev) {
    if (!isTooltipOpen()) return;
    var t = ev.target;
    if (t && typeof t.closest === 'function' && t.closest('.nextShiftTooltip')) return;
    hideTooltipNow();
  }, { passive: true, capture: true });

  document.querySelectorAll('.deptCard .empRow').forEach(bindEmployeeRow);
})();
"""


# =========================
# HTML Builders
# =========================
def dept_card_html(dept_name: str, dept_color: dict, buckets: dict, open_group: str = None) -> str:
    # buckets = {group_key: [{"name": ..., "shift": ...}, ...], ...}
    total = sum(len(buckets.get(k, [])) for k in GROUP_ORDER)
    if total == 0:
        return ""

    shifts_html = ""
    for group_key in GROUP_ORDER:
        emps = buckets.get(group_key, [])
        if not emps:
            continue

        # Determine shift display name (use English directly)
        if group_key == "Morning":
            display_name = "Morning"
        elif group_key == "Afternoon":
            display_name = "Afternoon"
        elif group_key == "Night":
            display_name = "Night"
        elif group_key == "Off Day":
            display_name = "Off Day"
        elif group_key == "Annual Leave":
            display_name = "Annual Leave"
        elif group_key == "Sick Leave":
           display_name = "Sick Leave"
        elif group_key == "Training":
            display_name = "Training"
        elif group_key == "Standby":
            display_name = "Standby"
        else:
            display_name = "Other"

        colors = SHIFT_COLORS.get(group_key, SHIFT_COLORS["Other"])
        count = len(emps)
        open_attr = ' open' if (group_key == open_group) else ''

        rows_html = ""
        for i, e in enumerate(emps):
            alt = " empRowAlt" if i % 2 == 1 else ""
            name_attr = html_escape(e["name"], quote=True)
            ar_name_attr = html_escape(name_i18n.arabic_display(e["name"]), quote=True)
            rows_html += f"""<div class="empRow{alt}" data-emp-name="{name_attr}" role="button" tabindex="0">
      <span class="empName" data-name-ar="{ar_name_attr}">{e['name']}</span>
       <span class="empStatus" style="color:{colors['status_color']};">{e['shift']}</span>
     </div>"""

        shifts_html += f"""
    <details class="shiftCard" data-shift="{group_key}" style="border:1px solid {colors['border']}; background:{colors['bg']}"{open_attr}>
      <summary class="shiftSummary" style="background:{colors['summary_bg']}; border-bottom:1px solid {colors['summary_border']};">
        <span class="shiftIcon">{colors['icon']}</span>
        <span class="shiftLabel" style="color:{colors['label_color']};">{display_name}</span>
        <span class="shiftCount" style="background:{colors['count_bg']}; color:{colors['count_color']};">{count}</span>
      </summary>
      <div class="shiftBody">
        {rows_html}
      </div>
    </details>
            """

    icon_svg = """
<svg width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.2" stroke-linecap="round" stroke-linejoin="round">
  <path d="M3 21h18M3 10h18M5 21V10l7-6 7 6v11"/>
  <rect x="9" y="14" width="2" height="3"/>
  <rect x="13" y="14" width="2" height="3"/>
</svg>
"""

    return f"""
    <div class="deptCard">
      <div style="height:5px; background:linear-gradient(to right, {dept_color['grad_from']}, {dept_color['grad_to']});"></div>

      <div class="deptHead" style="border-bottom:2px solid {dept_color['border']};">
        <div class="deptIcon" style="background:{dept_color['light']}; color:{dept_color['base']};">
          {icon_svg}
        </div>
        <div class="deptTitle">{dept_name}</div>
        <div class="deptBadge" style="background:{dept_color['light']}; color:{dept_color['base']}; border:1px solid {dept_color['border']};">
          <span style="font-size:10px;opacity:.7;display:block;margin-bottom:1px;text-transform:uppercase;letter-spacing:.5px;">Total</span>
          <span style="font-size:17px;font-weight:900;">{total}</span>
        </div>
      </div>

      <div class="shiftStack">
{shifts_html}
      </div>
    </div>
    """

def page_shell_html(date_label: str, iso_date: str, employees_total: int, departments_total: int,
                     dept_cards_html: str, cta_url: str, sent_time: str, source_name: str = "", last_updated: str = "", is_now_page: bool = False,
                     min_date: str = "", max_date: str = "", notice_html: str = "") -> str:

    pages_base = (PAGES_BASE_URL or infer_pages_base_url()).rstrip("/")
    min_attr = f'min="{min_date}"' if min_date else ""
    max_attr = f'max="{max_date}"' if max_date else ""

    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1, viewport-fit=cover">
  <meta name="x-apple-disable-message-reformatting">
  <script defer src="{pages_base}/ios-tap-fix.js?v={IOS_PERF_VER}"></script>
  <title>Duty Roster</title>
  <style>
    /* ═══════ RESET ═══════ */
    :root {{
      --safe-top: env(safe-area-inset-top, 0px);
      --safe-bottom: env(safe-area-inset-bottom, 0px);
    }}
    html, body {{
      width:100%;
      overflow-x:hidden;
    }}
    body {{
      margin:0; padding:0;
      min-height:100dvh;
      background:#eef1f7;
      font-family:'Segoe UI', system-ui, -apple-system, BlinkMacSystemFont, Roboto, Helvetica, Arial, sans-serif;
      color:#0f172a;
      -webkit-font-smoothing:antialiased;
      touch-action:manipulation;
    }}
    * {{ box-sizing:border-box; }}

    /* ═══════ WRAP ═══════ */
    .wrap {{ max-width:680px; margin:0 auto; padding:calc(16px + var(--safe-top)) 14px calc(28px + var(--safe-bottom)); }}

    /* ═══════ HEADER ═══════ */
    .header {{
      background:linear-gradient(135deg, #1e40af 0%, #1976d2 50%, #0ea5e9 100%);
      color:#fff;
      padding:26px 18px 24px;
      border-radius:20px;
      text-align:center;
      box-shadow:0 4px 16px rgba(30,64,175,.14);
      position:relative;
      overflow:hidden;
    }}
    .header::before {{
      content:''; position:absolute;
      top:-30px; right:-40px;
      width:140px; height:140px;
      border-radius:50%;
      background:rgba(255,255,255,.08);
    }}
    .header::after {{
      content:''; position:absolute;
      bottom:-50px; left:-30px;
      width:160px; height:160px;
      border-radius:50%;
      background:rgba(255,255,255,.06);
    }}
    .header .bannerTitle {{
      margin:0;
      position:relative;
      z-index:1;
      line-height:1.1;
      color:#fff;
    }}
    .header .bannerTitleEyebrow {{
      display:block;
      font-size:11px;
      font-weight:700;
      letter-spacing:.22em;
      text-transform:uppercase;
      opacity:.88;
      margin-bottom:5px;
    }}
    .header .bannerTitleMain {{
      display:block;
      font-size:28px;
      font-weight:800;
      letter-spacing:-.03em;
    }}
    body.ar .header .bannerTitleEyebrow {{
      letter-spacing:.06em;
      text-transform:none;
      font-size:12px;
    }}
    body.ar .header .bannerTitleMain {{
      font-size:26px;
      letter-spacing:0;
    }}

    /* زر اللغة */
{LANG_TOGGLE_CSS}
    body.ar {{ direction:rtl; font-family:'Segoe UI',Tahoma,Arial,sans-serif; }}
    .empRow, .empName, .empStatus {{ direction:ltr !important; unicode-bidi:embed; text-align:left !important; }}

    .welcomeChip {{
      display:none;
      text-decoration:none;
      cursor:pointer;
    }}
    .welcomeChip.visible {{
      display:flex;
    }}
    .welcomeChip .chipLabel {{
      max-width:88px;
      overflow:hidden;
      text-overflow:ellipsis;
    }}
    .waveHand {{
      display:inline-flex;
      align-items:center;
      justify-content:center;
      line-height:0;
      transform-origin:50% 88%;
      animation:waveHand 1.8s ease-in-out infinite;
    }}
    @keyframes waveHand {{
      0%, 50%, 100% {{ transform:rotate(0deg); }}
      10% {{ transform:rotate(16deg); }}
      20% {{ transform:rotate(-10deg); }}
      30% {{ transform:rotate(16deg); }}
      40% {{ transform:rotate(-6deg); }}
    }}

    /* Date Picker Wrapper */
    .datePickerWrapper {{
      position:relative;
      display:inline-block;
      margin-top:14px;
      z-index:20;
      min-height:44px;
      min-width:min(100%, 220px);
      touch-action:manipulation;
      -webkit-tap-highlight-color:transparent;
    }}
    .header .dateTag {{
      display:inline-flex;
      align-items:center;
      gap:8px;
      background:rgba(255,255,255,.18);
      padding:5px 18px;
      border-radius:10px;
      font-size:13px;
      font-weight:600;
      letter-spacing:.3px;
      cursor:pointer;
      transition:all .3s;
      border:2px solid rgba(255,255,255,.2);
      -webkit-tap-highlight-color:transparent;
      user-select:none;
      -webkit-user-select:none;
      direction:ltr;
      position:relative;
      z-index:3;
      pointer-events:auto;
      color:#fff;
      text-shadow:0 1px 2px rgba(0,0,0,.72),0 0 5px rgba(0,0,0,.38),0 0 1px rgba(255,255,255,.5);
    }}
    .dateTag-icon {{
      display:inline-flex;
      align-items:center;
      justify-content:center;
      flex-shrink:0;
      line-height:0;
      color:#fff;
      pointer-events:none;
    }}
    .dateTag-icon svg {{
      display:block;
      width:16px;
      height:16px;
      pointer-events:none;
      filter:drop-shadow(0 1px 1px rgba(0,0,0,.7)) drop-shadow(0 0 2px rgba(255,255,255,.45));
    }}
    .dateTag-label {{
      line-height:1.2;
      pointer-events:none;
      text-shadow:0 1px 2px rgba(0,0,0,.72),0 0 5px rgba(0,0,0,.38),0 0 1px rgba(255,255,255,.5);
    }}
    .header .dateTag:hover {{
      background:rgba(255,255,255,.25);
      transform:translateY(-1px);
    }}
    /* Transparent date input over #dateTag — native picker on iOS + desktop */
    .datePickerWrapper #datePicker {{
      position:absolute;
      inset:0;
      width:100%;
      height:100%;
      min-height:44px;
      margin:0;
      padding:0;
      opacity:0;
      cursor:pointer;
      font-size:16px;
      line-height:44px;
      border:none;
      z-index:5;
      pointer-events:auto;
      color:transparent;
      background:transparent;
      touch-action:manipulation;
    }}

    a.summaryChip, button.summaryChip, .langToggle, .roster-cta-btn, button.shiftFilterBtn {{
      touch-action:manipulation;
      -webkit-tap-highlight-color:transparent;
    }}

    .header::before,
    .header::after {{
      pointer-events:none;
    }}

    /* ═══════ SUMMARY BAR ═══════ */
    .summaryBar {{ 
      display:flex; 
      justify-content:center; 
      align-items:stretch;
      gap:12px; 
      margin-top:14px;
      flex-wrap:wrap;
      position:relative;
      z-index:30;
      isolation:isolate;
    }}
    .summaryBar a.summaryChip,
    .summaryBar button.summaryChip {{
      position:relative;
      z-index:1;
    }}
    .summaryBar a.summaryChip *,
    .summaryBar button.summaryChip *,
    .quickActions.roster-cta-btn .roster-cta-icon,
    .quickActions.roster-cta-btn .roster-cta-label {{
      pointer-events:none;
    }}
    .importBottom {{
      position:relative;
      z-index:25;
    }}
    a.summaryChip:hover {{
      transform:translateY(-3px);
      box-shadow:0 8px 20px rgba(15,23,42,.12);
    }}
    a.summaryChip.importChip .chipVal {{ color:#0ea5e9; }}
    a.summaryChip.importChip:hover {{ box-shadow:0 8px 20px rgba(14,165,233,.18); }}
    a.summaryChip.trainingChip .chipVal {{ color:#7c3aed; }}
    a.summaryChip.trainingChip:hover {{ box-shadow:0 8px 20px rgba(124,58,237,.18); }}
    a.summaryChip.diffChip .chipVal {{ color:#ef4444; }}
    a.summaryChip.diffChip:hover {{ box-shadow:0 8px 20px rgba(239,68,68,.18); }}
    .summaryChip {{
      background:#fff;
      border:1px solid rgba(15,23,42,.1);
      border-radius:14px;
      padding:10px 12px;
      text-align:center;
      box-shadow:0 2px 8px rgba(15,23,42,.06);
      transition:all .25s ease;
      min-width:72px;
      display:flex;
      flex-direction:column;
      align-items:center;
      justify-content:flex-start;
    }}
    .summaryChip .chipVal {{ font-size:22px; font-weight:900; color:#1e40af; height:26px; display:flex; align-items:center; justify-content:center; line-height:1; }}
{CHIP_ICON_CSS}
    #summarySwitchChip .chipVal {{ transition:opacity .2s ease; }}
    #summarySwitchChip .chipLabel {{ transition:opacity .2s ease; }}
    .summaryChip .chipLabel {{ 
      font-size:9.5px;
      font-weight:600; 
      color:#64748b; 
      text-transform:uppercase; 
      letter-spacing:.4px; 
      margin-top:4px;
      line-height:1.1;
      white-space:nowrap;
    }}

    /* ═══════ SHIFT FILTER BUTTONS AS CHIPS ═══════ */
    button.summaryChip.shiftFilterBtn {{
      border:2px solid transparent;
      position:relative;
      overflow:hidden;
      padding:10px 14px; /* padding أصغر للأزرار */
    }}
    button.summaryChip.shiftFilterBtn:hover {{
      transform:translateY(-3px);
      box-shadow:0 8px 20px rgba(15,23,42,.12);
    }}
    button.summaryChip.shiftFilterBtn:focus {{
      outline:none;
    }}
    button.summaryChip.shiftFilterBtn:focus:not(.active) {{
      border-color:transparent;
    }}
    button.summaryChip.shiftFilterBtn.active {{
      border-color:currentColor;
      box-shadow:0 6px 16px rgba(15,23,42,.18);
    }}
    button.summaryChip.shiftFilterBtn.active::before {{
      content:'';
      position:absolute;
      top:0;left:0;right:0;bottom:0;
      background:currentColor;
      opacity:.06;
    }}
    
    /* ألوان الورديات */
    button.shiftFilterBtn.morning {{
      color:#f59e0b;
    }}
    button.shiftFilterBtn.morning .chipVal {{ color:#f59e0b; }}
    button.shiftFilterBtn.morning .chipLabel {{ color:#92400e; }}
    
    button.shiftFilterBtn.afternoon {{
      color:#f97316;
    }}
    button.shiftFilterBtn.afternoon .chipVal {{ color:#f97316; }}
    button.shiftFilterBtn.afternoon .chipLabel {{ color:#9a3412; }}
    
    button.shiftFilterBtn.night {{
      color:#8b5cf6;
    }}
    button.shiftFilterBtn.night .chipVal {{ color:#8b5cf6; }}
    button.shiftFilterBtn.night .chipLabel {{ color:#5b21b6; }}
    
    button.shiftFilterBtn.all {{
      color:#1e40af;
    }}
    button.shiftFilterBtn.all .chipVal {{ color:#1e40af; }}
    button.shiftFilterBtn.all .chipLabel {{ color:#1e40af; }}
    
    /* للشاشات المتوسطة */
    @media (max-width:900px){{
      .summaryBar {{ flex-wrap:wrap; }} /* السماح بالانتقال للصف الثاني */
    }}
    
    /* للموبايل */
    @media (max-width:600px){{
      .summaryBar {{ gap:6px; }}
      .summaryChip {{ padding:8px 8px; min-width:60px; }}
      .summaryChip .chipVal {{ font-size:18px; }}
      .summaryChip .chipLabel {{ font-size:8.5px; letter-spacing:.2px; }}
    }}


    /* ═══════ DEPARTMENT CARD ═══════ */
    .deptCard {{
      margin-top:18px;
      background:#fff;
      border-radius:18px;
      overflow:hidden;
      border:1px solid rgba(15,23,42,.07);
      box-shadow:0 4px 18px rgba(15,23,42,.08);
    }}
{PERF_RENDER_CSS}
    .deptHead {{
      display:flex;
      align-items:center;
      gap:12px;
      padding:14px 16px;
      background:#fff;
      cursor:pointer;
    }}
    .deptHead::after {{
      content:'▾';
      color:#94a3b8;
      font-size:14px;
      margin-inline-start:6px;
      transition:transform .2s ease;
    }}
    .deptCard:not(.collapsed) .deptHead::after {{ transform:rotate(180deg); }}
    .deptIcon {{
      width:40px; height:40px;
      border-radius:12px;
      display:flex; align-items:center; justify-content:center;
      flex-shrink:0;
    }}
    .deptTitle {{ font-size:18px; font-weight:800; color:#1e293b; flex:1; letter-spacing:-.2px; }}
    .deptBadge {{ min-width:48px; padding:6px 10px; border-radius:12px; text-align:center; }}
    .deptCard.collapsed .deptHead {{ border-bottom:none !important; }}
    .deptCard.collapsed .shiftStack {{ display:none; }}

    /* ═══════ SHIFT STACK ═══════ */
    .shiftStack {{ padding:10px; display:flex; flex-direction:column; gap:8px; }}

    /* ═══════ SHIFT CARD — <details> ═══════ */
    .shiftCard {{
      border-radius:14px;
      overflow:hidden;
    }}

    .shiftSummary {{
      display:flex;
      align-items:center;
      gap:10px;
      padding:11px 14px;
      cursor:pointer;
      list-style:none;
      -webkit-appearance:none;
      appearance:none;
      user-select:none;
    }}
    .shiftSummary::-webkit-details-marker {{ display:none; }}
    .shiftSummary::marker              {{ display:none; }}

    .shiftIcon  {{ font-size:20px; line-height:1; flex-shrink:0; }}
    .shiftLabel {{ font-size:15px; font-weight:800; flex:1; letter-spacing:-.1px; }}
    .shiftCount {{
      font-size:13px; font-weight:800;
      padding:3px 10px; border-radius:20px;
      flex-shrink:0;
    }}

    /* chevron يدور لما يفتح */
    .shiftSummary::after {{
      content:'▾';
      font-size:14px;
      color:#94a3b8;
      transition:transform .2s;
      flex-shrink:0;
    }}
    .shiftCard[open] .shiftSummary::after {{
      transform:rotate(180deg);
    }}

    .shiftBody {{ background:rgba(255,255,255,.7); }}

    /* ── employee row (tap/long-press targets whole row in JS) ── */
    .empRow {{
      display:flex;
      align-items:center;
      justify-content:space-between;
      padding:9px 16px;
      border-top:1px solid rgba(15,23,42,.06);
      cursor:pointer;
      -webkit-user-select:none;
      user-select:none;
      -webkit-tap-highlight-color:transparent;
    }}
    .empRowAlt {{ background:rgba(15,23,42,.02); }}
    .empName  {{
      font-size:15px; font-weight:700; color:#1e293b;
      -webkit-user-select:none;
      user-select:none;
      pointer-events:none;
    }}
    .empStatus {{ font-size:13px; font-weight:600; }}
    .nextShiftTooltip {{
      position: fixed;
      z-index: 12000;
      min-width: 250px;
      max-width: 320px;
      background: #ffffff;
      border: 1px solid rgba(148, 163, 184, 0.35);
      border-radius: 14px;
      box-shadow: 0 14px 34px rgba(15, 23, 42, 0.20);
      overflow: hidden;
      pointer-events: none;
      opacity: 0;
      transform: translateY(4px);
      transition: opacity .16s ease, transform .16s ease;
    }}
    .nextShiftTooltip.show {{
      opacity: 1;
      transform: translateY(0);
      pointer-events: auto;
    }}
    .nextShiftHead {{
      padding: 8px 10px;
      background: linear-gradient(135deg, #2563eb, #1d4ed8);
      border-bottom: 1px solid rgba(15, 23, 42, 0.12);
      color: #fff;
      display: flex;
      align-items: flex-start;
      justify-content: space-between;
      gap: 8px;
    }}
    .nextShiftHeadText {{ flex: 1; min-width: 0; }}
    .nextShiftClose {{
      flex: 0 0 auto;
      width: 26px;
      height: 26px;
      border: none;
      border-radius: 8px;
      background: rgba(255, 255, 255, 0.2);
      color: #fff;
      font-size: 18px;
      line-height: 1;
      cursor: pointer;
      display: grid;
      place-items: center;
      padding: 0;
    }}
    .nextShiftClose:hover {{ background: rgba(255, 255, 255, 0.32); }}
    .nextShiftEmp {{
      font-size: 12px; font-weight: 800; line-height: 1.2;
      white-space: nowrap; overflow: hidden; text-overflow: ellipsis;
    }}
    .nextShiftTitle {{
      font-size: 10px; font-weight: 700; opacity: .9; margin-top: 2px;
    }}
    .nextShiftBody {{
      padding: 8px 9px 9px; background: rgba(255, 255, 255, 0.96);
    }}
    .nextShiftItem {{
      margin-top: 6px;
      padding: 6px 8px;
      border-radius: 10px;
      background: rgba(241, 245, 249, 0.92);
      display: flex;
      flex-direction: row;
      align-items: center;
      gap: 8px;
    }}
    .nextShiftDay {{
      font-size: 11px; font-weight: 800; color: #334155;
      line-height: 1.2; min-width: 36px;
    }}
    .nextShiftDate {{
      font-size: 10px; color: #64748b; line-height: 1.2; min-width: 58px;
    }}
    .nextShiftCode {{
      margin-inline-start: auto;
      font-size: 12px; font-weight: 800; color: #0f172a; line-height: 1.2;
    }}
    .nextShiftEmpty {{
      font-size: 11px;
      color: #64748b;
      margin-top: 4px;
    }}

    /* ═══════ QUICK ACTIONS — see scripts/roster_cta_snippets.py ═══════ */
    .quickActions.roster-cta {{
      --cta-font: "Segoe UI", system-ui, -apple-system, sans-serif;
      --cta-gap: 10px;
      margin-top: 22px;
      padding: 0 2px;
      display: grid;
      grid-template-columns: repeat(6, 1fr);
      gap: var(--cta-gap);
      width: 100%;
      max-width: 100%;
      margin-inline: auto;
    }}
    .quickActions.roster-cta:not(.roster-cta--import) > .roster-cta-btn:nth-child(1) {{
      grid-column: 1 / span 2;
    }}
    .quickActions.roster-cta:not(.roster-cta--import) > .roster-cta-btn:nth-child(2) {{
      grid-column: 3 / span 2;
    }}
    .quickActions.roster-cta:not(.roster-cta--import) > .roster-cta-btn:nth-child(3) {{
      grid-column: 5 / span 2;
    }}
    .quickActions.roster-cta:not(.roster-cta--import) > .roster-cta-btn--share {{
      grid-column: 2 / span 2;
    }}
    .quickActions.roster-cta:not(.roster-cta--import) > .roster-cta-btn--apps {{
      grid-column: 4 / span 2;
    }}
    .roster-cta-btn {{
      display: inline-flex;
      align-items: center;
      justify-content: center;
      gap: 8px;
      min-height: 44px;
      padding: 10px 12px;
      border-radius: 999px;
      border: 1.5px solid transparent;
      font-family: var(--cta-font);
      font-size: 13px;
      font-weight: 700;
      line-height: 1.2;
      text-decoration: none;
      cursor: pointer;
      touch-action: manipulation;
      -webkit-tap-highlight-color: transparent;
      box-shadow: none;
      transition: transform 0.15s ease, box-shadow 0.15s ease, background 0.15s ease, border-color 0.15s ease;
    }}
    button.roster-cta-btn {{
      appearance: none;
      -webkit-appearance: none;
      font: inherit;
    }}
    .roster-cta-icon {{
      flex-shrink: 0;
      display: inline-flex;
      align-items: center;
      justify-content: center;
      width: 18px;
      height: 18px;
      line-height: 0;
    }}
    .roster-cta-icon svg {{ display: block; width: 18px; height: 18px; }}
    .roster-cta-label {{
      white-space: nowrap;
      overflow: hidden;
      text-overflow: ellipsis;
    }}
    .roster-cta-btn--roster {{
      background: #e8f1ff;
      border-color: #b8c9f5;
      color: #1e3a8a;
    }}
    .roster-cta-btn--subscribe {{
      background: #ffffff;
      border-color: #d4c4f7;
      color: #1f2937;
    }}
    .roster-cta-btn--compare {{
      background: #fffbeb;
      border-color: #fcd34d;
      color: #1f2937;
    }}
    .roster-cta-btn--muted {{
      background: #f1f5f9;
      border-color: #cbd5e1;
      color: #475569;
    }}
    .roster-cta-btn--share {{
      background: #ecfdf5;
      border-color: #86efac;
      color: #166534;
    }}
    .roster-cta-btn--apps {{
      background: #f0f9ff;
      border-color: #7dd3fc;
      color: #0369a1;
    }}
    .roster-cta-btn--texture {{
      background: #f5f3ff;
      border-color: #c4b5fd;
      color: #5b21b6;
    }}
    .roster-cta--import {{
      grid-template-columns: 1fr 1fr;
    }}
    @media (hover: hover) {{
      .roster-cta-btn:hover {{
        transform: translateY(-1px);
        box-shadow: 0 4px 12px rgba(15, 23, 42, 0.08);
      }}
      .roster-cta-btn--roster:hover {{ background: #dce8ff; }}
      .roster-cta-btn--subscribe:hover {{ background: #faf5ff; }}
      .roster-cta-btn--compare:hover {{ background: #fef3c7; }}
      .roster-cta-btn--share:hover {{ background: #d1fae5; }}
      .roster-cta-btn--apps:hover {{ background: #e0f2fe; }}
      .roster-cta-btn--texture:hover {{ background: #ede9fe; }}
      .roster-cta-btn--muted:hover {{ background: #e2e8f0; }}
    }}
    .roster-cta-btn:active {{
      transform: translateY(0) scale(0.98);
      box-shadow: none;
    }}
    .roster-cta-btn:focus-visible {{
      outline: 2px solid rgba(37, 99, 235, 0.45);
      outline-offset: 2px;
    }}
    @media (max-width: 380px) {{
      .roster-cta-btn {{
        padding: 9px 8px;
        font-size: 11px;
        gap: 5px;
        min-height: 40px;
      }}
      .roster-cta-icon {{ font-size: 16px; }}
    }}

    /* ═══════ SITE SHARE MODAL ═══════ */
    .siteShareSheet {{
      position:fixed; inset:0; display:none; align-items:center; justify-content:center;
      background:rgba(15,23,42,.45); z-index:10001; padding:16px;
      pointer-events:none; visibility:hidden;
    }}
    .siteShareSheet.open {{ display:flex; pointer-events:auto; visibility:visible; }}
    .siteShareCard {{
      width:min(100%, 360px); background:#fff; border-radius:18px; padding:18px 16px 14px;
      border:1px solid rgba(15,23,42,.1); box-shadow:0 20px 48px rgba(15,23,42,.22);
      text-align:center;
    }}
    .siteShareTitle {{ font-size:17px; font-weight:800; color:#0f172a; margin:0 0 4px; }}
    .siteShareHint {{ font-size:12px; color:#64748b; margin:0 0 14px; line-height:1.4; }}
    .siteShareQr {{
      display:flex; align-items:center; justify-content:center;
      min-height:220px; margin:0 auto 12px;
      background:#f8fafc; border-radius:14px; border:1px solid #e2e8f0; padding:10px;
    }}
    .siteShareUrl {{
      font-size:11px; color:#475569; word-break:break-all; line-height:1.45;
      margin:0 0 14px; padding:8px 10px; background:#f1f5f9; border-radius:10px;
    }}
    .siteShareActions {{
      display:grid; grid-template-columns:1fr 1fr; gap:10px; margin-bottom:10px;
    }}
    .siteShareActions .roster-cta-btn--compare {{ grid-column:1 / -1; }}
    .siteShareCloseWrap {{ margin-top:4px; }}
    .siteShareCloseWrap .roster-cta-btn {{ width:100%; }}

    /* ═══════ RELATED APPS MODAL ═══════ */
    .siteAppsSheet {{
      position:fixed; inset:0; display:none; align-items:center; justify-content:center;
      background:rgba(15,23,42,.45); z-index:10002; padding:16px;
      pointer-events:none; visibility:hidden;
    }}
    .siteAppsSheet.open {{ display:flex; pointer-events:auto; visibility:visible; }}
    .siteAppsCard {{
      width:min(100%,400px); max-height:min(92vh,560px); overflow:auto;
      -webkit-overflow-scrolling:touch; background:#fff; border-radius:18px;
      padding:18px 14px 14px; border:1px solid rgba(15,23,42,.1);
      box-shadow:0 20px 48px rgba(15,23,42,.22); text-align:center;
    }}
    .siteAppsTitle {{ font-size:17px; font-weight:800; color:#0f172a; margin:0 0 4px; }}
    .siteAppsHint {{ font-size:12px; color:#64748b; margin:0 0 14px; line-height:1.4; }}
    .siteAppsGrid {{ display:grid; grid-template-columns:1fr 1fr; gap:10px; margin-bottom:12px; text-align:start; }}
    .siteAppsLink {{
      display:flex; flex-direction:column; align-items:center; justify-content:center; gap:8px;
      min-height:96px; padding:12px 10px; border-radius:14px; border:1px solid #e2e8f0;
      background:#f8fafc; text-decoration:none; color:#0f172a;
      -webkit-tap-highlight-color:transparent;
    }}
    .siteAppsLink-icon {{
      display:flex; align-items:center; justify-content:center; width:44px; height:44px;
      border-radius:12px; background:#fff; border:1px solid #e2e8f0;
      box-shadow:0 2px 8px rgba(15,23,42,.06);
    }}
    .siteAppsLink-title {{ font-size:12px; font-weight:800; line-height:1.25; text-align:center; color:#0f172a; }}
    .siteAppsLink-sub {{ font-size:10px; font-weight:600; color:#64748b; line-height:1.3; text-align:center; }}
    .siteAppsLink--games {{
      grid-column:1 / -1; flex-direction:row; min-height:72px; justify-content:flex-start;
      padding-inline:14px; gap:12px;
    }}
    .siteAppsLink--games .siteAppsLink-text {{ display:flex; flex-direction:column; align-items:flex-start; gap:2px; flex:1; }}
    .siteAppsLink--games .siteAppsLink-title, .siteAppsLink--games .siteAppsLink-sub {{ text-align:start; }}
    .siteAppsCloseWrap {{ margin-top:4px; }}
    .siteAppsCloseWrap .roster-cta-btn {{ width:100%; }}

{SHIFT_COPY_CSS}
    /* ═══════ FOOTER ═══════ */
    .footer {{ margin-top:18px; text-align:center; font-size:12px; color:#94a3b8; padding:12px 0; line-height:1.9; }}
    .footer strong {{ color:#64748b; }}

    /* ═══════ SHARE/SAVE CAPTURE SHEET ═══════ */
    .captureSheet {{
      position:fixed; inset:0; display:none; align-items:center; justify-content:center;
      background:rgba(15,23,42,.45); z-index:9999; padding:12px 10px;
      pointer-events:none; visibility:hidden;
    }}
    .captureSheet.open {{ display:flex; pointer-events:auto; visibility:visible; }}
    .captureSheetCard {{
      width:min(100%,420px); max-height:min(92dvh, 900px);
      display:flex; flex-direction:column; overflow:hidden;
      background:#fff; border-radius:16px; padding:12px;
      border:1px solid rgba(15,23,42,.1); box-shadow:0 16px 40px rgba(15,23,42,.24);
    }}
    .captureSheetTitle {{
      flex-shrink:0; font-size:13px; font-weight:800; color:#334155; padding:4px 6px 10px;
    }}
    .capturePreviewWrap {{
      flex:1 1 auto; min-height:0;
      max-height:min(58dvh, 520px); overflow-y:auto; overflow-x:hidden;
      margin:0 0 10px; border-radius:12px; border:1px solid #e2e8f0; background:#f8fafc;
      -webkit-overflow-scrolling:touch;
    }}
    .capturePreviewImg {{
      display:block; width:100%; max-width:100%; height:auto; margin:0;
      border:none; border-radius:0; background:transparent;
    }}
    .captureSheetActions {{
      flex-shrink:0; display:grid; grid-template-columns:1fr 1fr; gap:8px;
    }}
    .captureSheetBtn {{
      border:none; border-radius:12px; padding:11px 10px; cursor:pointer; font:800 12px/1 'Segoe UI',sans-serif;
    }}
    .captureShareBtn {{ background:linear-gradient(135deg,#1e40af,#1976d2); color:#fff; }}
    .captureSaveBtn {{ background:#e8eefc; color:#1e40af; }}
    .captureCancelBtn {{
      flex-shrink:0; margin-top:8px; width:100%; background:#f1f5f9; color:#475569;
    }}
    .captureBusy {{
      position:fixed; top:12px; left:50%; transform:translateX(-50%);
      background:#0f172a; color:#fff; font-size:12px; font-weight:700;
      border-radius:999px; padding:8px 12px; z-index:10000; display:none;
      pointer-events:none;
    }}
    .captureBusy.open {{ display:block; }}

    /* ═══════ MOBILE ═══════ */
    @media (max-width:480px){{
      .wrap            {{ padding:12px 10px 22px; }}
      .deptTitle       {{ font-size:16px; }}
      .empName         {{ font-size:14px; }}
      .empStatus       {{ font-size:12px; }}
      .shiftLabel      {{ font-size:14px; }}
      .summaryBar      {{ gap:8px; }}
      .summaryChip     {{ padding:8px 14px; }}
      .summaryChip .chipVal {{ font-size:19px; }}
    }}

  </style>{ROSTER_PWA_HEAD_SNIPPET}
</head>
<body>
<div class="wrap">

  <!-- ════ HEADER ════ -->
  <div class="header">
    {LANG_TOGGLE_HTML}
    <h1 id="pageTitle" class="bannerTitle">
      <span class="bannerTitleEyebrow" id="pageTitleEyebrow">Export</span>
      <span class="bannerTitleMain" id="pageTitleMain">Duty Roster</span>
    </h1>
    <div class="datePickerWrapper">
      <label class="dateTag" id="dateTag" for="datePicker"><span class="dateTag-icon" aria-hidden="true"><svg viewBox="0 0 24 24" width="16" height="16" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" aria-hidden="true"><rect x="3" y="4" width="18" height="18" rx="2"/><path d="M16 2v4M8 2v4M3 10h18"/></svg></span><span class="dateTag-label" id="dateTagLabel">{date_label}</span></label>
      <input id="datePicker" type="date" value="{iso_date}" {min_attr} {max_attr} aria-label="Select roster date" />
    </div>
  </div>

  {notice_html if notice_html else ""}

  <!-- ════ SUMMARY CHIPS ════ -->
  <div class="summaryBar">
    <div class="summaryChip" id="summarySwitchChip">
      <div class="chipVal" id="summarySwitchVal">{employees_total}</div>
      <div class="chipLabel" id="summarySwitchLabel" data-key="employees">Employees</div>
    </div>
    <a href="{pages_base}/my-schedules/index.html" id="myScheduleBtn" class="summaryChip" style="text-decoration:none;">
      {CHIP_SCHEDULE_HTML}
      <div class="chipLabel" data-key="mySchedule">My Schedule</div>
    </a>
    <a href="{pages_base}/import/" id="importBtn" class="summaryChip importChip" style="text-decoration:none;">
      {CHIP_FLIGHT_HTML}
      <div class="chipLabel" data-key="importRoster">Import</div>
    </a>
    <a href="{pages_base}/my-schedules/index.html" id="welcomeChip" class="summaryChip welcomeChip" title="Go to your schedule" style="text-decoration:none;">
      {CHIP_WAVE_HTML}
      <div class="chipLabel" id="welcomeName"></div>
    </a>
    <a href="{pages_base}/training/" id="trainingBtn" class="summaryChip trainingChip" style="text-decoration:none;">
      {CHIP_TRAINING_HTML}
      <div class="chipLabel" data-key="trainingPage">Training</div>
    </a>
    <a href="{pages_base}/roster-diff/index.html" id="diffChipBtn" class="summaryChip diffChip" style="text-decoration:none;">
      {CHIP_DIFF_HTML}
      <div class="chipLabel" data-key="diffPage">Diff</div>
    </a>
    {"" if not is_now_page else f'''
    <button class="summaryChip shiftFilterBtn morning" data-shift="Morning" style="cursor:pointer;">
      {CHIP_MORNING_HTML}
      <div class="chipLabel" data-key="morning">Morning</div>
    </button>
    <button class="summaryChip shiftFilterBtn afternoon" data-shift="Afternoon" style="cursor:pointer;">
      {CHIP_AFTERNOON_HTML}
      <div class="chipLabel" data-key="afternoon">Afternoon</div>
    </button>
    <button class="summaryChip shiftFilterBtn night" data-shift="Night" style="cursor:pointer;">
      {CHIP_NIGHT_HTML}
      <div class="chipLabel" data-key="night">Night</div>
    </button>
    <button class="summaryChip shiftFilterBtn all" data-shift="All" style="cursor:pointer;">
      {CHIP_ALL_HTML}
      <div class="chipLabel" data-key="allShifts">All Shifts</div>
    </button>
    '''}
  </div>

  <!-- ════ DEPARTMENT CARDS ════ -->
  {dept_cards_html}

  <!-- ════ CTA ════ -->
{export_cta_html()}

  <!-- ════ COPY SHIFT ════ -->
{SHIFT_COPY_BUTTON_HTML}

  <!-- ════ FOOTER ════ -->
  <div class="footer">
    <strong style="color:#475569;font-size:13px;">Last Updated:</strong> <strong style="color:#1e40af;" id="siteLastUpdated" data-site-last-updated="1">{last_updated}</strong><br>
    <strong style="color:#475569;font-size:13px;">Source:</strong> <strong>{source_name}</strong>
  </div>

</div>

{SITE_SHARE_MODAL_HTML}

{SITE_APPS_MODAL_HTML}

{SHIFT_COPY_MODAL_HTML}

<div id="captureBusy" class="captureBusy">Preparing image...</div>
<div id="captureSheet" class="captureSheet" aria-hidden="true">
  <div class="captureSheetCard">
    <div class="captureSheetTitle">Share or save image</div>
    <div class="capturePreviewWrap">
      <img id="capturePreview" class="capturePreviewImg" alt="Snapshot preview" />
    </div>
    <div class="captureSheetActions">
      <button id="captureShareBtn" class="captureSheetBtn captureShareBtn" type="button">Share</button>
      <button id="captureSaveBtn" class="captureSheetBtn captureSaveBtn" type="button">Save</button>
    </div>
    <button id="captureCancelBtn" class="captureSheetBtn captureCancelBtn" type="button">Cancel</button>
  </div>
</div>

<script>
var __html2canvasLoading = null;
function ensureHtml2Canvas() {{
  if (typeof html2canvas === 'function') return Promise.resolve(html2canvas);
  if (__html2canvasLoading) return __html2canvasLoading;
  __html2canvasLoading = new Promise(function(resolve, reject) {{
    var s = document.createElement('script');
    s.src = 'https://cdn.jsdelivr.net/npm/html2canvas@1.4.1/dist/html2canvas.min.js';
    s.async = true;
    s.onload = function() {{
      if (typeof html2canvas === 'function') resolve(html2canvas);
      else reject(new Error('html2canvas missing'));
    }};
    s.onerror = function() {{
      __html2canvasLoading = null;
      reject(new Error('html2canvas load failed'));
    }};
    document.head.appendChild(s);
  }});
  return __html2canvasLoading;
}}
</script>
<script>
function getSiteRootPath() {{
  if (location.protocol === 'file:') return '';
  var path = location.pathname || '/';
  if (path.includes('/roster-site/')) return '/roster-site';
  if (location.hostname && location.hostname.endsWith('github.io')) {{
    var segs = path.split('/').filter(Boolean);
    if (segs.length >= 2 && segs[1] === 'docs') return '/' + segs[0] + '/docs';
    return segs.length ? '/' + segs[0] : '';
  }}
  return '';
}}

function getSiteRootUrl() {{
  return location.origin + getSiteRootPath();
}}

function setCaptureBusy(isBusy) {{
  var busy = document.getElementById('captureBusy');
  if(!busy) return;
  busy.classList.toggle('open', !!isBusy);
}}

function saveBlobFile(blob, fileName) {{
  var url = URL.createObjectURL(blob);
  var a = document.createElement('a');
  a.href = url;
  a.download = fileName;
  document.body.appendChild(a);
  a.click();
  a.remove();
  setTimeout(function(){{ URL.revokeObjectURL(url); }}, 1200);
}}

function openCaptureSheet(blob, fileName, captureMode) {{
  var sheet = document.getElementById('captureSheet');
  var shareBtn = document.getElementById('captureShareBtn');
  var saveBtn = document.getElementById('captureSaveBtn');
  var cancelBtn = document.getElementById('captureCancelBtn');
  var preview = document.getElementById('capturePreview');
  var title = sheet ? sheet.querySelector('.captureSheetTitle') : null;
  if(!sheet || !shareBtn || !saveBtn || !cancelBtn) return;
  var modeText = captureMode || 'UNKNOWN';
  if (title) title.textContent = 'Share or save image (' + modeText + ')';

  var file = new File([blob], fileName, {{ type:'image/png' }});
  var shareSupported = !!(navigator.share && navigator.canShare && navigator.canShare({{ files:[file] }}));
  shareBtn.style.display = shareSupported ? '' : 'none';

  var previewUrl = null;
  if (preview) {{
    var previewWrap = preview.closest('.capturePreviewWrap');
    if (!previewWrap && preview.parentNode) {{
      previewWrap = document.createElement('div');
      previewWrap.className = 'capturePreviewWrap';
      preview.parentNode.insertBefore(previewWrap, preview);
      previewWrap.appendChild(preview);
    }}
    if (preview.dataset.revokeUrl) {{
      try {{ URL.revokeObjectURL(preview.dataset.revokeUrl); }} catch(e) {{}}
      preview.dataset.revokeUrl = '';
    }}
    previewUrl = URL.createObjectURL(blob);
    preview.src = previewUrl;
    preview.dataset.revokeUrl = previewUrl;
    preview.style.display = '';
    if (previewWrap) previewWrap.scrollTop = 0;
  }}

  function closeSheet() {{
    sheet.classList.remove('open');
    sheet.setAttribute('aria-hidden', 'true');
    shareBtn.onclick = null;
    saveBtn.onclick = null;
    cancelBtn.onclick = null;
    sheet.onclick = null;
    if (preview && preview.dataset.revokeUrl) {{
      try {{ URL.revokeObjectURL(preview.dataset.revokeUrl); }} catch(e) {{}}
      preview.dataset.revokeUrl = '';
      preview.removeAttribute('src');
    }}
  }}

  saveBtn.onclick = function() {{
    saveBlobFile(blob, fileName);
    closeSheet();
  }};
  shareBtn.onclick = async function() {{
    try {{
      await navigator.share({{ files:[file], title:'Duty Roster Snapshot' }});
    }} catch(e) {{
      // User may cancel share dialog.
    }}
    closeSheet();
  }};
  cancelBtn.onclick = closeSheet;
  sheet.onclick = function(e) {{ if(e.target === sheet) closeSheet(); }};

  sheet.classList.add('open');
  sheet.setAttribute('aria-hidden', 'false');
}}

function rosterSnapshotLayoutWidth() {{
  var iw = window.innerWidth || 400;
  /* Same visual column as phone: narrow on desktop, full usable width on small screens (max ~430). */
  return Math.min(Math.max(iw - 28, 280), 430);
}}

/** Gradient strip + .deptHead (everything before .shiftStack) for shift snapshots. */
function buildDeptBannerForSnapshot(deptCard) {{
  if (!deptCard) return null;
  var stack = deptCard.querySelector('.shiftStack');
  if (!stack) return null;
  var banner = document.createElement('div');
  var el = deptCard.firstElementChild;
  while (el && el !== stack) {{
    banner.appendChild(el.cloneNode(true));
    el = el.nextElementSibling;
  }}
  if (!banner.children.length) return null;
  banner.style.marginBottom = '8px';
  banner.style.width = '100%';
  banner.style.boxSizing = 'border-box';
  return banner;
}}

function openAllShiftsOnDept(dept) {{
  var saved = [];
  if (!dept) return function() {{}};
  var wasCollapsed = dept.classList.contains('collapsed');
  saved.push({{ kind: 'collapsed', el: dept, value: wasCollapsed }});
  dept.classList.remove('collapsed');
  var stack = dept.querySelector('.shiftStack');
  if (stack) {{
    saved.push({{ kind: 'stackDisplay', el: stack, value: stack.style.display }});
    stack.style.display = 'flex';
  }}
  dept.querySelectorAll('details.shiftCard').forEach(function(d) {{
    saved.push({{ kind: 'shiftOpen', el: d, value: d.open }});
    d.open = true;
    d.setAttribute('open', '');
  }});
  return function restoreDeptShiftState() {{
    for (var i = saved.length - 1; i >= 0; i--) {{
      var s = saved[i];
      if (s.kind === 'shiftOpen') {{
        s.el.open = s.value;
        if (!s.value) s.el.removeAttribute('open');
      }} else if (s.kind === 'stackDisplay') {{
        s.el.style.display = s.value;
      }} else if (s.kind === 'collapsed' && s.value) {{
        s.el.classList.add('collapsed');
      }}
    }}
  }};
}}

function prepareSnapshotClone(root) {{
  if (!root) return;
  root.classList.remove('collapsed');
  root.style.contentVisibility = 'visible';
  root.style.contain = 'none';
  root.style.overflow = 'visible';
  root.style.height = 'auto';
  root.style.maxHeight = 'none';
  root.querySelectorAll('.shiftStack').forEach(function(stack) {{
    stack.style.display = 'flex';
    stack.style.overflow = 'visible';
    stack.style.height = 'auto';
  }});
  root.querySelectorAll('details.shiftCard').forEach(function(card) {{
    card.open = true;
    card.setAttribute('open', '');
    card.style.display = 'block';
    card.style.overflow = 'visible';
    card.style.height = 'auto';
    card.style.maxHeight = 'none';
  }});
  root.querySelectorAll('.shiftBody').forEach(function(body) {{
    body.style.display = 'block';
    body.style.overflow = 'visible';
    body.style.height = 'auto';
    body.style.maxHeight = 'none';
  }});
  root.querySelectorAll('.empRow').forEach(function(row) {{
    row.style.display = 'flex';
  }});
}}

function getBannerBackgroundForSnapshot() {{
  var header = document.querySelector('.header');
  if (header) {{
    var cs = window.getComputedStyle(header);
    var bgImg = cs.backgroundImage;
    if (bgImg && bgImg !== 'none') {{
      return {{
        image: bgImg,
        size: cs.backgroundSize || 'cover',
        position: cs.backgroundPosition || '62% center',
        repeat: cs.backgroundRepeat || 'no-repeat',
        color: cs.backgroundColor || ''
      }};
    }}
  }}
  try {{
    var bn = localStorage.getItem('roster_banner_choice');
    if (bn && /^banner\\d+\\.jpg$/i.test(bn)) {{
      var url = getSiteRootUrl() + '/assets/banners/' + bn;
      return {{ image: 'url("' + url + '")', size: 'cover', position: '62% center', repeat: 'no-repeat', color: '' }};
    }}
  }} catch(e) {{}}
  return {{
    image: 'linear-gradient(135deg, #1e40af 0%, #1976d2 50%, #0ea5e9 100%)',
    size: 'cover',
    position: 'center',
    repeat: 'no-repeat',
    color: ''
  }};
}}

function buildBannerHeaderForSnapshot() {{
  var liveHeader = document.querySelector('.header');
  if (!liveHeader) return null;
  var clone = liveHeader.cloneNode(true);
  clone.className = 'header captureBannerHeader';
  clone.querySelectorAll('.langToggle, #banner-changer-btn, #datePicker').forEach(function(el) {{
    el.remove();
  }});
  var bg = getBannerBackgroundForSnapshot();
  clone.style.position = 'relative';
  clone.style.overflow = 'hidden';
  clone.style.marginBottom = '6px';
  clone.style.borderRadius = '20px';
  clone.style.padding = '22px 16px 20px';
  clone.style.textAlign = 'center';
  clone.style.color = '#fff';
  clone.style.minHeight = '112px';
  clone.style.boxSizing = 'border-box';
  clone.style.boxShadow = '0 4px 16px rgba(30,64,175,.14)';
  clone.style.backgroundImage = bg.image;
  clone.style.backgroundSize = bg.size;
  clone.style.backgroundPosition = bg.position;
  clone.style.backgroundRepeat = bg.repeat;
  if (bg.color) clone.style.backgroundColor = bg.color;
  clone.querySelectorAll('.bannerTitle, .bannerTitleEyebrow, .bannerTitleMain, .dateTag, .dateTag-label').forEach(function(el) {{
    el.style.position = 'relative';
    el.style.zIndex = '2';
    el.style.color = '#fff';
    el.style.textShadow = '0 1px 3px rgba(0,0,0,.55)';
  }});
  var picker = clone.querySelector('.datePickerWrapper');
  if (picker) {{
    picker.style.position = 'relative';
    picker.style.zIndex = '2';
    picker.style.marginTop = '10px';
    picker.style.display = 'inline-block';
  }}
  return clone;
}}

function injectCaptureSnapshotStyles(wrap) {{
  if (!wrap || wrap.querySelector('[data-capture-style="1"]')) return;
  var style = document.createElement('style');
  style.setAttribute('data-capture-style', '1');
  style.textContent = [
    '[data-capture-wrap]{{box-sizing:border-box;}}',
    '[data-capture-wrap] .captureBannerHeader{{width:100%;margin-bottom:10px;}}',
    '[data-capture-wrap] .deptCard{{overflow:visible!important;content-visibility:visible!important;contain:none!important;border-radius:18px!important;box-shadow:0 4px 18px rgba(15,23,42,.08)!important;margin:0!important;}}',
    '[data-capture-wrap] .deptHead::after,[data-capture-wrap] .shiftSummary::after{{display:none!important;}}',
    '[data-capture-wrap] .shiftStack{{display:flex!important;flex-direction:column!important;gap:6px!important;padding:8px 10px!important;}}',
    '[data-capture-wrap] .shiftCardFlat,[data-capture-wrap] .shiftCard{{display:block!important;overflow:visible!important;margin:0!important;border-radius:14px!important;}}',
    '[data-capture-wrap] .shiftBody{{display:block!important;overflow:visible!important;height:auto!important;}}',
    '[data-capture-wrap] .empRow{{display:flex!important;align-items:center!important;padding:8px 14px!important;min-height:36px!important;}}'
  ].join('');
  wrap.insertBefore(style, wrap.firstChild);
}}

function flattenShiftCardsInClone(root) {{
  if (!root) return;
  root.querySelectorAll('details.shiftCard').forEach(function(details) {{
    var flat = document.createElement('div');
    flat.className = details.className + ' shiftCardFlat';
    if (details.getAttribute('style')) flat.setAttribute('style', details.getAttribute('style'));
    flat.style.display = 'block';
    flat.style.overflow = 'visible';
    flat.style.height = 'auto';
    flat.style.maxHeight = 'none';
    Array.from(details.children).forEach(function(ch) {{
      var tag = String(ch.tagName || '').toUpperCase();
      if (tag === 'SUMMARY') {{
        var sumDiv = document.createElement('div');
        sumDiv.className = ch.className;
        if (ch.getAttribute('style')) sumDiv.setAttribute('style', ch.getAttribute('style'));
        sumDiv.innerHTML = ch.innerHTML;
        flat.appendChild(sumDiv);
      }} else {{
        var cloned = ch.cloneNode(true);
        cloned.style.display = 'block';
        cloned.style.overflow = 'visible';
        cloned.style.height = 'auto';
        cloned.style.maxHeight = 'none';
        if (cloned.classList && cloned.classList.contains('shiftBody')) {{
          cloned.querySelectorAll('.empRow').forEach(function(row) {{ row.style.display = 'flex'; }});
        }}
        flat.appendChild(cloned);
      }}
    }});
    details.replaceWith(flat);
  }});
}}

function measureCaptureWrapHeight(wrap) {{
  if (!wrap) return 400;
  var total = 0;
  Array.from(wrap.children).forEach(function(child) {{
    if (child.getAttribute && child.getAttribute('data-capture-style') === '1') return;
    var rect = child.getBoundingClientRect();
    if (rect.height <= 0) return;
    var st = window.getComputedStyle(child);
    var mb = parseFloat(st.marginBottom) || 0;
    var mt = parseFloat(st.marginTop) || 0;
    total += rect.height + mb + mt;
  }});
  var wrapSt = window.getComputedStyle(wrap);
  var pad = (parseFloat(wrapSt.paddingTop) || 0) + (parseFloat(wrapSt.paddingBottom) || 0);
  var sum = Math.ceil(total + pad);
  if (sum > 80) return sum + 2;
  return Math.ceil(wrap.scrollHeight || wrap.offsetHeight || 0) + 2;
}}

async function captureRosterElement(target, fileNamePrefix, opts) {{
  opts = opts || {{}};
  if(!target) return;
  try {{
    await ensureHtml2Canvas();
  }} catch (e) {{
    return;
  }}
  if(typeof html2canvas !== 'function') return;
  setCaptureBusy(true);
  var restoreLiveDept = null;
  var wrap = null;
  try {{
    function waitForCaptureLayout() {{
      return new Promise(function(resolve) {{
        if (typeof requestAnimationFrame === 'function') {{
          requestAnimationFrame(function() {{
            requestAnimationFrame(resolve);
          }});
        }} else {{
          setTimeout(resolve, 32);
        }}
      }});
    }}
    var isDepartment = !!opts.expandAllShifts;
    var header = document.querySelector('.header');
    var layoutW = rosterSnapshotLayoutWidth();
    wrap = document.createElement('div');
    wrap.setAttribute('data-capture-wrap', '1');
    wrap.style.position = 'absolute';
    wrap.style.left = '-9999px';
    wrap.style.top = '0';
    wrap.style.width = layoutW + 'px';
    wrap.style.boxSizing = 'border-box';
    wrap.style.background = '#eef1f7';
    wrap.style.padding = isDepartment ? '10px 12px 6px' : '14px';
    wrap.style.visibility = 'visible';
    wrap.style.opacity = '1';
    wrap.style.pointerEvents = 'none';
    wrap.style.overflow = 'visible';
    wrap.style.height = 'auto';
    wrap.style.maxHeight = 'none';

    if (isDepartment) {{
      var bannerHeader = buildBannerHeaderForSnapshot();
      if (bannerHeader) wrap.appendChild(bannerHeader);
    }} else if (header) {{
      var headerClone = header.cloneNode(true);
      headerClone.style.marginBottom = '10px';
      wrap.appendChild(headerClone);
    }}
    if (opts.prependClone && opts.prependClone.nodeType === 1) {{
      var pre = opts.prependClone.cloneNode(true);
      pre.style.marginBottom = '8px';
      wrap.appendChild(pre);
    }}

    var sourceTarget = target;
    if (opts.deptCaptureId) {{
      var byId = document.querySelector('.deptCard[data-dept-capture-id="' + opts.deptCaptureId + '"]');
      if (byId) sourceTarget = byId;
    }}
    if (isDepartment && target && typeof target.closest === 'function') {{
      var sourceDept = target.closest('.deptCard');
      if (sourceDept) sourceTarget = sourceDept;
    }}

    var targetClone = null;
    if (isDepartment) {{
      var dept = sourceTarget;
      restoreLiveDept = openAllShiftsOnDept(dept);
      await waitForCaptureLayout();
      targetClone = dept.cloneNode(true);
      restoreLiveDept();
      restoreLiveDept = null;
      prepareSnapshotClone(targetClone);
      flattenShiftCardsInClone(targetClone);
    }} else {{
      targetClone = sourceTarget.cloneNode(true);
      if (targetClone.classList && targetClone.classList.contains('shiftCard')) {{
        targetClone.open = true;
        targetClone.setAttribute('open', '');
      }}
      prepareSnapshotClone(targetClone);
    }}

    targetClone.style.marginTop = '0';
    targetClone.style.width = '100%';
    targetClone.style.maxWidth = '100%';
    targetClone.style.boxSizing = 'border-box';
    wrap.appendChild(targetClone);
    injectCaptureSnapshotStyles(wrap);
    document.body.appendChild(wrap);
    await waitForCaptureLayout();
    await waitForCaptureLayout();

    var captureHeight = 0;
    if (isDepartment) {{
      captureHeight = measureCaptureWrapHeight(wrap);
    }}

    var canvasOpts = {{
      backgroundColor: '#eef1f7',
      scale: Math.max(2, window.devicePixelRatio || 1),
      useCORS: true,
      logging: false,
      scrollX: 0,
      scrollY: -window.scrollY,
      ignoreElements: function(el) {{
        return el && el.id === 'captureSheet';
      }},
      onclone: function(doc) {{
        var clonedWrap = doc.querySelector('[data-capture-wrap="1"]');
        if (!clonedWrap) return;
        clonedWrap.style.left = '0';
        clonedWrap.style.visibility = 'visible';
        clonedWrap.style.opacity = '1';
        clonedWrap.style.overflow = 'visible';
        clonedWrap.style.height = 'auto';
        clonedWrap.style.maxHeight = 'none';
        clonedWrap.querySelectorAll('.deptCard').forEach(function(dept) {{
          prepareSnapshotClone(dept);
          flattenShiftCardsInClone(dept);
        }});
        clonedWrap.querySelectorAll('.shiftCard').forEach(function(card) {{
          prepareSnapshotClone(card);
        }});
      }}
    }};
    var scale = Math.max(2, window.devicePixelRatio || 1);
    if (isDepartment && captureHeight > 14000) scale = 1.5;
    if (isDepartment && captureHeight > 22000) scale = 1;
    canvasOpts.scale = scale;

    var canvas = null;
    try {{
      canvas = await html2canvas(wrap, canvasOpts);
    }} catch (captureErr) {{
      var fallbackOpts = Object.assign({{}}, canvasOpts, {{
        scale: 1,
        height: undefined,
        windowHeight: undefined
      }});
      canvas = await html2canvas(wrap, fallbackOpts);
    }}
    if (wrap && wrap.parentNode) wrap.remove();
    wrap = null;

    canvas.toBlob(function(blob){{
      if(!blob) return;
      var stamp = new Date().toISOString().slice(0,16).replace(/[:T]/g,'-');
      var mode = isDepartment ? 'DEPARTMENT' : 'SHIFT';
      console.log('[capture] mode=' + mode + ' canvas=' + canvas.width + 'x' + canvas.height);
      openCaptureSheet(blob, fileNamePrefix + '-' + stamp + '.png', mode);
    }}, 'image/png');
  }} catch(err) {{
    console.error('[capture]', err);
    alert('Could not create image snapshot.');
  }} finally {{
    if (restoreLiveDept) restoreLiveDept();
    if (wrap && wrap.parentNode) wrap.remove();
    setCaptureBusy(false);
  }}
}}

function goToEmployeeSchedule(empName) {{
  var match = empName.match(/-\s*(\d{{3,}})/);
  var base = getSiteRootUrl() + '/my-schedules/index.html';

  if (match) {{
    location.href = base + '?emp=' + encodeURIComponent(match[1]);
  }} else {{
    location.href = base + '?name=' + encodeURIComponent(empName);
  }}
}}

(function(){{
  var picker = document.getElementById('datePicker');
  if(!picker) return;

  function getMuscatTodayIso() {{
    var now = new Date();
    var muscatTime = new Date(now.getTime() + (4 * 60 * 60 * 1000) + (now.getTimezoneOffset() * 60 * 1000));
    return muscatTime.getFullYear() + '-' +
      String(muscatTime.getMonth() + 1).padStart(2, '0') + '-' +
      String(muscatTime.getDate()).padStart(2, '0');
  }}

  function formatIsoLabel(iso) {{
    var d = new Date(iso + 'T00:00:00');
    if (isNaN(d.getTime())) return iso;
    return d.toLocaleDateString('en-GB', {{ day: 'numeric', month: 'long', year: 'numeric' }});
  }}

  function syncHeaderDate(iso) {{
    var tag = document.getElementById('dateTag');
    if (tag) {{
      var dateLbl = document.getElementById('dateTagLabel');
      var dateText = formatIsoLabel(iso);
      if (dateLbl) dateLbl.textContent = dateText;
      else tag.textContent = dateText;
    }}
  }}

  var path = window.location.pathname || '/';
  var pageDateMatch = path.match(/\/date\/(\d{{4}})-(\d{{2}})-(\d{{2}})\//);
  var effectiveIso = pageDateMatch
    ? (pageDateMatch[1] + '-' + pageDateMatch[2] + '-' + pageDateMatch[3])
    : getMuscatTodayIso();
  picker.value = effectiveIso;
  syncHeaderDate(effectiveIso);

  var DATE_PICKER_BUSY_KEY = 'rosterDatePickerBusy';

  function setDatePickerBusy(on) {{
    try {{
      if (on) sessionStorage.setItem(DATE_PICKER_BUSY_KEY, '1');
      else sessionStorage.removeItem(DATE_PICKER_BUSY_KEY);
    }} catch (e) {{}}
  }}

  window.openDatePicker = function() {{
    if (!picker) return;
    setDatePickerBusy(true);
    try {{ picker.focus({{ preventScroll: true }}); }} catch (e) {{ picker.focus(); }}
    if (typeof picker.showPicker === 'function') {{
      try {{ picker.showPicker(); return; }} catch (e) {{}}
    }}
    try {{ picker.click(); }} catch (e2) {{}}
  }};

  function onDateWrapActivate(e) {{
    if (e) {{
      e.preventDefault();
      e.stopPropagation();
    }}
    openDatePicker();
  }}


  var dateTagEl = document.getElementById('dateTag');
  if (dateTagEl) {{
    dateTagEl.addEventListener('click', function(e) {{
      if (e) e.preventDefault();
      openDatePicker();
    }});
  }}

  var dateWrap = document.querySelector('.datePickerWrapper');
  if (dateWrap) {{
    dateWrap.addEventListener('touchend', onDateWrapActivate, {{ passive: false }});
    dateWrap.addEventListener('click', function(e) {{
      if (e.target === picker) return;
      onDateWrapActivate(e);
    }});
    function onPickerActivate(e) {{
      if (e) {{
        e.preventDefault();
        e.stopPropagation();
      }}
      setDatePickerBusy(true);
      openDatePicker();
    }}
    picker.addEventListener('click', onPickerActivate);
    picker.addEventListener('keydown', function(e) {{
      if (e.key === 'Enter' || e.key === ' ') {{
        onPickerActivate(e);
      }}
    }});
  }}
  picker.addEventListener('focus', function() {{ setDatePickerBusy(true); }});
  picker.addEventListener('blur', function() {{
    setTimeout(function() {{ setDatePickerBusy(false); }}, 400);
  }});

  var USER_DATE_NAV_KEY = 'rosterUserPickedDate';

  function consumeUserDateNavigation() {{
    if (sessionStorage.getItem(USER_DATE_NAV_KEY) !== '1') return false;
    sessionStorage.removeItem(USER_DATE_NAV_KEY);
    return true;
  }}

  function markUserDateNavigation() {{
    sessionStorage.setItem(USER_DATE_NAV_KEY, '1');
  }}

  function normalizePathname(p) {{
    return (p || '/')
      .replace(/\/date\/\\d{{4}}-\\d{{2}}-\\d{{2}}\\/.*$/i, '/')
      .replace(/\/import\/date\/\\d{{4}}-\\d{{2}}-\\d{{2}}\\/.*$/i, '/')
      .replace(/\/import\/\\d{{4}}-\\d{{2}}-\\d{{2}}\\/.*$/i, '/')
      .replace(/\/\\d{{4}}-\\d{{2}}-\\d{{2}}\\/.*$/i, '/')
      .replace(/\/now\\/.*$/i, '/')
      .replace(/\/index\\.html$/i, '')
      .replace(/\\/+$/, '');
  }}

  function isImportRosterPath(p) {{
    return /\/import(\/|$)/i.test(p || '');
  }}

  function buildDateBasePath() {{
    var root = typeof getSiteRootPath === 'function' ? getSiteRootPath() : '';
    var p = path || location.pathname || '';
    if (isImportRosterPath(p)) {{
      var base = root || normalizePathname(p);
      if (!base || base === '/') {{
        return '/import';
      }}
      if (!/\\/import$/i.test(base)) {{
        base = base.replace(/\\/+$/, '') + '/import';
      }}
      return base;
    }}
    if (root) return root;
    return normalizePathname(p);
  }}

  function redirectToDate(iso, isNowPage) {{
    window.location.replace(buildDateBasePath() + '/date/' + iso + '/' + (isNowPage ? 'now/' : ''));
  }}

  // ═══════════════════════════════════════════════════
  // التحقق من التاريخ وإعادة التوجيه للـ today
  // ═══════════════════════════════════════════════════
  function checkAndRedirectToToday() {{
    var isNowPage = path.includes('/now');
    var todayIso = getMuscatTodayIso();

    if (!path.includes('/date/')) {{
      window.location.replace(buildDateBasePath() + '/date/' + todayIso + '/' + (isNowPage ? 'now/' : ''));
      return true;
    }}

    var dateMatch = path.match(/\/date\/(\\d{{4}})-(\\d{{2}})-(\\d{{2}})\//);
    if (!dateMatch) return false;

    var pageIso = dateMatch[1] + '-' + dateMatch[2] + '-' + dateMatch[3];
    if (pageIso === todayIso) {{
      sessionStorage.removeItem(USER_DATE_NAV_KEY);
      sessionStorage.removeItem('pageLoaded');
      return false;
    }}

    if (consumeUserDateNavigation()) return false;

    redirectToDate(todayIso, isNowPage);
    return true;
  }}

  function resyncTodayIfNeeded() {{
    if (!path.includes('/date/')) return;
    try {{
      if (sessionStorage.getItem(DATE_PICKER_BUSY_KEY) === '1') return;
    }} catch (e) {{}}
    var dateMatch = path.match(/\/date\/(\\d{{4}})-(\\d{{2}})-(\\d{{2}})\//);
    if (!dateMatch) return;
    var pageIso = dateMatch[1] + '-' + dateMatch[2] + '-' + dateMatch[3];
    var todayIso = getMuscatTodayIso();
    if (pageIso !== todayIso && sessionStorage.getItem(USER_DATE_NAV_KEY) !== '1') {{
      redirectToDate(todayIso, path.includes('/now'));
    }}
  }}

  if (checkAndRedirectToToday()) return;

  window.addEventListener('pageshow', function(ev) {{
    if (ev.persisted) resyncTodayIfNeeded();
  }});
  document.addEventListener('visibilitychange', function() {{
    if (document.visibilityState === 'visible') resyncTodayIfNeeded();
  }});

  // ═══════════════════════════════════════════════════
  // عند تغيير التاريخ → انتقل للصفحة المناسبة
  // ═══════════════════════════════════════════════════
  picker.addEventListener('change', function() {{
    if (!picker.value) return;

    markUserDateNavigation();
    sessionStorage.removeItem('pageLoaded');

    var isNowPage = (window.location.pathname || '').includes('/now');
    var base = buildDateBasePath();
    var target = base + '/date/' + picker.value + '/';
    if (isNowPage) target += 'now/';

    window.location.href = target;
  }});

}})();

// ═══════════════════════════════════════════════════
// Long-press capture for section / shift
// ═══════════════════════════════════════════════════
(function(){{
  var LONG_PRESS_MS = 550;
  var suppressNextClick = false;

  function bindLongPress(el, onLongPress){{
    if(!el) return;
    var timer = null;
    var moved = false;
    var startX = 0;
    var startY = 0;
    var tracking = false;
    var MOVE_TOLERANCE_PX = 8;

    function clear(){{
      if(timer) clearTimeout(timer);
      timer = null;
    }}

    el.addEventListener('pointerdown', function(e){{
      if(e.button !== 0 && e.button !== undefined) return;
      moved = false;
      tracking = true;
      startX = Number(e.clientX || 0);
      startY = Number(e.clientY || 0);
      clear();
      timer = setTimeout(function(){{
        timer = null;
        if(!tracking || moved) return;
        suppressNextClick = true;
        onLongPress(e);
      }}, LONG_PRESS_MS);
    }});
    el.addEventListener('pointermove', function(e){{
      if(!tracking) return;
      var dx = Math.abs(Number(e.clientX || 0) - startX);
      var dy = Math.abs(Number(e.clientY || 0) - startY);
      if(dx > MOVE_TOLERANCE_PX || dy > MOVE_TOLERANCE_PX){{
        moved = true;
        clear();
      }}
    }});
    el.addEventListener('pointerup', function(){{ tracking = false; clear(); }});
    el.addEventListener('pointercancel', function(){{ tracking = false; clear(); }});
    el.addEventListener('pointerleave', function(e){{
      if(e.pointerType === 'mouse' || moved){{
        tracking = false;
        clear();
      }}
    }});
    el.addEventListener('click', function(e){{
      if(suppressNextClick){{
        e.preventDefault();
        e.stopPropagation();
        if (typeof e.stopImmediatePropagation === 'function') e.stopImmediatePropagation();
        setTimeout(function() {{ suppressNextClick = false; }}, 0);
      }}
    }}, true);
    el.addEventListener('contextmenu', function(e){{
      if(suppressNextClick) e.preventDefault();
    }});
  }}

  document.querySelectorAll('.deptHead').forEach(function(head){{
    if (head.dataset.deptCaptureBound !== '1') {{
      head.dataset.deptCaptureBound = '1';
      bindLongPress(head, function(){{
        var cardForLongPress = head.closest('.deptCard');
        if(!cardForLongPress) return;
        cardForLongPress.classList.remove('collapsed');
        cardForLongPress.querySelectorAll('details.shiftCard').forEach(function(shiftCard){{
          shiftCard.open = true;
          shiftCard.setAttribute('open', '');
        }});
        captureRosterElement(cardForLongPress, 'department', {{ expandAllShifts: true }});
      }});
    }}

    if (head.dataset.deptShiftToggleBound === '1') return;
    head.dataset.deptShiftToggleBound = '1';
    head.addEventListener('click', function() {{
      if (suppressNextClick) return;
      var card = head.closest('.deptCard');
      if (!card) return;
      card.classList.remove('collapsed');
      var shifts = card.querySelectorAll('details.shiftCard');
      if (!shifts.length) return;
      var step = parseInt(card.dataset.deptShiftStep || '0', 10);
      if (isNaN(step) || step < 0 || step > 2) step = 0;
      if (step === 0) {{
        shifts.forEach(function(d) {{ d.setAttribute('open', ''); }});
        card.dataset.deptShiftStep = '1';
      }} else if (step === 1) {{
        shifts.forEach(function(d) {{ d.removeAttribute('open'); }});
        card.dataset.deptShiftStep = '2';
      }} else {{
        var now = new Date();
        var muscatTime = new Date(now.getTime() + (4 * 60 * 60 * 1000) + (now.getTimezoneOffset() * 60 * 1000));
        var hour = muscatTime.getHours();
        var minute = muscatTime.getMinutes();
        var tt = hour * 60 + minute;
        var currentShift = (tt >= 21 * 60 || tt < 5 * 60) ? 'Night' : (tt >= 13 * 60 ? 'Afternoon' : 'Morning');
        shifts.forEach(function(d) {{ d.removeAttribute('open'); }});
        var target = null;
        shifts.forEach(function(d) {{
          if (d.dataset.shift === currentShift) target = d;
        }});
        if (!target) target = shifts[0];
        if (target) target.setAttribute('open', '');
        card.dataset.deptShiftStep = '0';
      }}
    }});
  }});

  document.querySelectorAll('.shiftSummary').forEach(function(summary){{
    bindLongPress(summary, function(){{
      var shiftCard = summary.closest('.shiftCard');
      if(!shiftCard) return;
      shiftCard.setAttribute('open', '');
      var deptCard = shiftCard.closest('.deptCard');
      var deptBanner = buildDeptBannerForSnapshot(deptCard);
      captureRosterElement(shiftCard, 'shift', deptBanner ? {{ prependClone: deptBanner }} : {{}});
    }});
  }});
}})();

// ══════════════════════════════════════════════════
// Language Toggle
// ══════════════════════════════════════════════════
var LANG = localStorage.getItem('rosterLang') || 'en';
window.__summaryCounts = window.__summaryCounts || {{ employees: {employees_total}, departments: {departments_total} }};
window.__summarySwitchMode = 'employees';
var T = {{
  en: {{
    titleEyebrow:'Export', titleMain:'Duty Roster', langBtn:'ع',
    employees:'Emp.', departments:'Depts.', total:'Total',
    morning:'Morning', afternoon:'Afternoon', night:'Night',
    offday:'Off Day', annualLeave:'Annual Leave', sickLeave:'Sick Leave',
    training:'Training', standby:'Standby', other:'Other',
    from:'FROM', to:'TO',
    viewFull:'Full Roster', subscribe:'Subscribe', compare:'Compare', shareSite:'Share Site', moreApps:'Apps',
    officers:'Officers', supervisors:'Supervisors', loadControl:'Load Control',
    exportChecker:'Export Checker', exportOps:'Export Operators', unassigned:'Unassigned',
    morning2:'Morning', afternoon2:'Afternoon', night2:'Night', allShifts:'All Shifts', mySchedule:'Schedule', importRoster:'Import', trainingPage:'Training', diffPage:'Diff',
    copyShift:'Copy Shift', copyTitle:'On-duty list', copyHint:'Copy or share a shift as WhatsApp text', copyDone:'Copied', copyEmpty:'No employees in this shift', copyFail:'Copy failed — long-press to copy', copyClose:'Close', copyAction:'Copy', shareAction:'Share', shareDone:'Shared',
  }},
  ar: {{
    titleEyebrow:'الصادر', titleMain:'جدول المناوبات', langBtn:'EN',
    employees:'الموظفون', departments:'الأقسام', total:'المجموع',
    morning:'صباح', afternoon:'ظهر', night:'ليل',
    offday:'إجازة', annualLeave:'إجازة سنوية', sickLeave:'إجازة مرضية',
    training:'تدريب', standby:'احتياط', other:'أخرى',
    from:'من', to:'إلى',
    viewFull:'الجدول الكامل', subscribe:'اشتراك', compare:'مقارنة', shareSite:'مشاركة الموقع', moreApps:'تطبيقات',
    officers:'الضباط', supervisors:'المشرفون', loadControl:'مراقبة الحمولة',
    exportChecker:'مدقق الصادرات', exportOps:'مشغلو الصادرات', unassigned:'غير مُعيَّن',
    morning2:'صباح', afternoon2:'ظهر', night2:'ليل', allShifts:'الكل', mySchedule:'جدولي', importRoster:'الوارد', trainingPage:'تدريب', diffPage:'فروقات',
    copyShift:'نسخ المناوبة', copyTitle:'قائمة المناوبين', copyHint:'انسخ أو شارك المناوبة كنص واتساب', copyDone:'تم نسخ', copyEmpty:'لا يوجد موظفون في هذه المناوبة', copyFail:'فشل النسخ — اضغط مطولاً للنسخ', copyClose:'إغلاق', copyAction:'نسخ', shareAction:'مشاركة', shareDone:'تمت المشاركة',
  }}
}};

function updateSummarySwitchChip() {{
  var val = document.getElementById('summarySwitchVal');
  var lbl = document.getElementById('summarySwitchLabel');
  if(!val || !lbl) return;
  var t = T[LANG] || T.en;
  var mode = window.__summarySwitchMode || 'employees';
  var counts = window.__summaryCounts || {{}};
  if(mode === 'departments') {{
    val.style.color = '#059669';
    val.textContent = counts.departments != null ? counts.departments : {departments_total};
    lbl.textContent = t.departments;
    lbl.dataset.key = 'departments';
  }} else {{
    val.style.color = '';
    val.textContent = counts.employees != null ? counts.employees : {employees_total};
    lbl.textContent = t.employees;
    lbl.dataset.key = 'employees';
  }}
}}

function startSummarySwitchLoop() {{
  if(window.__summarySwitchTimer) return;
  window.__summarySwitchTimer = setInterval(function(){{
    window.__summarySwitchMode = (window.__summarySwitchMode === 'employees') ? 'departments' : 'employees';
    updateSummarySwitchChip();
  }}, 2200);
}}

function applyLang(lang) {{
  var t=T[lang], isAr=lang==='ar';
  document.body.classList.toggle('ar',isAr);
  document.documentElement.setAttribute('lang',lang);
  var eyebrow=document.getElementById('pageTitleEyebrow');
  var main=document.getElementById('pageTitleMain');
  if(eyebrow) eyebrow.textContent=t.titleEyebrow;
  if(main) main.textContent=t.titleMain;
{APPLY_LANG_LANG_BTN_NEW.rstrip()}
  document.querySelectorAll('.chipLabel').forEach(function(el) {{
    var k=el.dataset.key;
    if(k==='employees') el.textContent=t.employees;
    else if(k==='departments') el.textContent=t.departments;
    else if(k==='morning') el.textContent=t.morning2;
    else if(k==='afternoon') el.textContent=t.afternoon2;
    else if(k==='night') el.textContent=t.night2;
    else if(k==='allShifts') el.textContent=t.allShifts;
    else if(k==='mySchedule') el.textContent=t.mySchedule;
    else if(k==='importRoster') el.textContent=t.importRoster;
    else if(k==='trainingPage') el.textContent=t.trainingPage;
    else if(k==='diffPage') el.textContent=t.diffPage;
  }});
  document.querySelectorAll('.deptBadge span:first-child').forEach(function(el) {{ el.textContent=t.total; }});
  var deptMap={{'Officers':t.officers,'Supervisors':t.supervisors,'Load Control':t.loadControl,
    'Export Checker':t.exportChecker,'Export Operators':t.exportOps,'Unassigned':t.unassigned}};
  document.querySelectorAll('.deptTitle').forEach(function(el) {{
    if(!el.dataset.key) el.dataset.key=el.textContent.trim();
    if(deptMap[el.dataset.key]) el.textContent=deptMap[el.dataset.key];
  }});
  var shiftMap={{'Morning':t.morning,'Afternoon':t.afternoon,'Night':t.night,
    'Off Day':t.offday,'Annual Leave':t.annualLeave,'Sick Leave':t.sickLeave,
    'Training':t.training,'Standby':t.standby,'Other':t.other}};
  document.querySelectorAll('.shiftLabel').forEach(function(el) {{
    if(!el.dataset.key) el.dataset.key=el.textContent.trim();
    if(shiftMap[el.dataset.key]) el.textContent=shiftMap[el.dataset.key];
  }});
  document.querySelectorAll('.empStatus .shiftRangeLabel, .empStatus span').forEach(function(el) {{
    var txt=el.textContent.trim();
    if(txt==='FROM'||txt==='من') el.textContent=t.from;
    if(txt==='TO'||txt==='إلى') el.textContent=t.to;
  }});
  document.querySelectorAll('.empName').forEach(function(el) {{
    if(el.dataset.nameEn===undefined) el.dataset.nameEn=el.textContent;
    var ar=el.getAttribute('data-name-ar');
    el.textContent=(isAr && ar) ? ar : el.dataset.nameEn;
  }});
  function setCtaLabel(id, text) {{
    var el = document.getElementById(id);
    if (!el) return;
    var lbl = el.querySelector('.roster-cta-label');
    if (lbl) lbl.textContent = text;
    else el.textContent = text;
  }}
  setCtaLabel('ctaBtn', t.viewFull);
  setCtaLabel('subscribeBtn', t.subscribe);
  setCtaLabel('compareBtn', t.compare);
  setCtaLabel('shareSiteBtn', t.shareSite);
  setCtaLabel('moreAppsBtn', t.moreApps);
  setCtaLabel('copyShiftBtn', t.copyShift);
  if(window.rosterSiteShare && window.rosterSiteShare.setLang) window.rosterSiteShare.setLang();
  if(window.rosterSiteApps && window.rosterSiteApps.setLang) window.rosterSiteApps.setLang();
  if(window.rosterSiteShiftCopy && window.rosterSiteShiftCopy.setLang) window.rosterSiteShiftCopy.setLang();
  var footer=document.querySelector('.footer');
  if(footer) {{
    var h=footer.innerHTML;
    if(isAr) {{
      h=h.replace('Last Updated','آخر تحديث'); h=h.replace('Source:','المصدر:');
    }} else {{
      h=h.replace('آخر تحديث','Last Updated'); h=h.replace('المصدر:','Source:');
    }}
    footer.innerHTML=h;
  }}
  localStorage.setItem('rosterLang',lang);
  LANG=lang;
  updateSummarySwitchChip();
}}
function toggleLang() {{ applyLang(LANG==='en'?'ar':'en'); }}

function setLocalCtaLinks() {{
  var root = getSiteRootPath();
  var c1 = document.getElementById('ctaBtn');
  var c2 = document.getElementById('subscribeBtn');
  if (c1) c1.href = root + '/now/';
  if (c2) c2.href = root + '/subscribe/';
}}

function setSummaryChipHrefs() {{
  var base = getSiteRootUrl();
  var my = document.getElementById('myScheduleBtn');
  var imp = document.getElementById('importBtn');
  var trn = document.getElementById('trainingBtn');
  var diff = document.getElementById('diffChipBtn');
  var welcome = document.getElementById('welcomeChip');
  if (my) my.href = base + '/my-schedules/index.html';
  if (imp) imp.href = base + '/import/';
  if (trn) trn.href = base + '/training/';
  if (diff) diff.href = base + '/roster-diff/index.html';
  if (welcome) {{
    var wid = localStorage.getItem('exportSavedEmpId') || localStorage.getItem('savedEmpId');
    var wbase = base + '/my-schedules/index.html';
    welcome.href = wid ? wbase + '?emp=' + encodeURIComponent(wid) : wbase;
  }}
}}
function goToTraining(e) {{
  if (e) e.preventDefault();
  var root = getSiteRootPath();
  location.href = root + '/training/';
}}
setLocalCtaLinks();
setSummaryChipHrefs();
applyLang(LANG);
startSummarySwitchLoop();

(function bindSummarySwitchScroll() {{
  var chip = document.getElementById('summarySwitchChip');
  if (!chip || chip.__scrollBound) return;
  chip.__scrollBound = true;
  chip.style.cursor = 'pointer';
  chip.setAttribute('role', 'button');
  chip.setAttribute('tabindex', '0');
  function ensureShuffleButton() {{
    if (document.getElementById('bgTextureShuffleBtn')) return;
    try {{
      var root = getSiteRootUrl();
      var src = root + '/bg-texture-shuffle.js?v={IOS_PERF_VER}';
      if (document.querySelector('script[data-local-src="' + src + '"]')) return;
      var s = document.createElement('script');
      s.src = src;
      s.defer = true;
      s.setAttribute('data-local-src', src);
      document.body.appendChild(s);
    }} catch (e) {{}}
  }}
  function scrollToBottom() {{
    ensureShuffleButton();
    function go() {{
      var root = document.scrollingElement || document.documentElement;
      var top = Math.max(0, root.scrollHeight - root.clientHeight);
      window.scrollTo({{ top: top, left: 0, behavior: 'smooth' }});
    }}
    go();
    // The "Shuffle background" button is injected lazily (requestIdleCallback,
    // up to ~3s), which grows the footer. Re-scroll several times so we always
    // land on the true bottom once that button appears.
    var delays = [150, 400, 800, 1400, 2200, 3200];
    delays.forEach(function (ms) {{ window.setTimeout(go, ms); }});
  }}
  chip.addEventListener('click', scrollToBottom);
  chip.addEventListener('keydown', function (e) {{
    if (e.key === 'Enter' || e.key === ' ') {{ e.preventDefault(); scrollToBottom(); }}
  }});
}})();

// ═══════════════════════════════════════════════════
// Department layout: all sections open + time-based shift open
// ═══════════════════════════════════════════════════
(function(){{
  var deptCards = Array.from(document.querySelectorAll('.deptCard'));
  if(!deptCards.length) return;
  var savedEmpId = localStorage.getItem('exportSavedEmpId') || localStorage.getItem('savedEmpId') || '';
  var savedEmpName = (localStorage.getItem('exportSavedEmpName') || localStorage.getItem('savedEmpName') || '').trim().toLowerCase();

  function getCurrentShiftForMuscat(){{
    var now = new Date();
    var muscatTime = new Date(now.getTime() + (4 * 60 * 60 * 1000) + (now.getTimezoneOffset() * 60 * 1000));
    var hour = muscatTime.getHours();
    var minute = muscatTime.getMinutes();
    var t = hour * 60 + minute;
    if(t >= 21 * 60 || t < 5 * 60) return 'Night';
    if(t >= 13 * 60) return 'Afternoon';
    return 'Morning';
  }}

  function openMatchingShift(card){{
    var shiftCards = Array.from(card.querySelectorAll('.shiftCard'));
    if(!shiftCards.length) return;
    shiftCards.forEach(function(shiftCard){{
      shiftCard.removeAttribute('open');
    }});
    var currentShift = getCurrentShiftForMuscat();
    var target = shiftCards.find(function(shiftCard){{
      return shiftCard.dataset.shift === currentShift;
    }}) || shiftCards[0];
    if(target) target.setAttribute('open', '');
  }}

  function cardHasSavedEmployee(card){{
    var rows = Array.from(card.querySelectorAll('.empName'));
    if(!rows.length) return false;
    return rows.some(function(row){{
      var txt = (row.textContent || '').trim();
      if(!txt) return false;
      if(savedEmpId && txt.indexOf(savedEmpId) !== -1) return true;
      if(savedEmpName && txt.toLowerCase().indexOf(savedEmpName) !== -1) return true;
      return false;
    }});
  }}

  // Keep all department sections visible/open.
  deptCards.forEach(function(card){{
    card.classList.remove('collapsed');
    openMatchingShift(card);
  }});

  // If a saved employee exists, move their department to top.
  if(savedEmpId || savedEmpName){{
    var targetCard = deptCards.find(cardHasSavedEmployee);
    if(targetCard && targetCard.parentElement){{
      var firstDeptCard = deptCards[0];
      if(firstDeptCard && firstDeptCard !== targetCard){{
        targetCard.parentElement.insertBefore(targetCard, firstDeptCard);
      }}
      openMatchingShift(targetCard);
    }}
  }}
}})();

// ═══════════════════════════════════════════════════
// Shift Filter (NOW PAGE ONLY) — runs after dept layout
// ═══════════════════════════════════════════════════
(function(){{
  var filterBtns = document.querySelectorAll('.shiftFilterBtn');
  if(!filterBtns.length) return;

  var allShiftCards = document.querySelectorAll('.shiftCard');

  function normalizeShift(raw){{
    if(!raw) return '';
    var t = String(raw).trim().toLowerCase();
    if(t === 'all' || t.includes('all shifts') || t === 'الكل') return 'All';
    if(t.includes('morning') || t.includes('صباح')) return 'Morning';
    if(t.includes('afternoon') || t.includes('ظهر') || t.includes('مساء')) return 'Afternoon';
    if(t.includes('night') || t.includes('ليل')) return 'Night';
    if(t.includes('off day') || (t.includes('off') && !t.includes('officer')) || t.includes('راحة') || t.includes('أوف')) return 'Off Day';
    if(t.includes('annual') || t.includes('سنوية')) return 'Annual Leave';
    if(t.includes('sick') || t.includes('مرض')) return 'Sick Leave';
    if(t.includes('training') || t.includes('تدريب')) return 'Training';
    if(t.includes('standby') || t.includes('ستاند') || t.includes('احتياط')) return 'Standby';
    if(t.includes('other') || t.includes('أخرى') || t.includes('اخرى')) return 'Other';
    return String(raw).trim();
  }}

  function btnShift(btn){{
    return normalizeShift(btn.getAttribute('data-shift') || btn.dataset.shift || '');
  }}

  filterBtns.forEach(function(btn){{
    var s = btnShift(btn);
    if(s) btn.setAttribute('data-shift', s);
  }});

  var shiftGroups = {{}};
  allShiftCards.forEach(function(card){{
    var label = card.querySelector('.shiftSummary .shiftLabel');
    var shiftType = normalizeShift(
      label ? (label.dataset.key || label.textContent) : (card.getAttribute('data-shift') || card.dataset.shift || '')
    );
    if(!shiftType) return;
    if(!shiftGroups[shiftType]) shiftGroups[shiftType] = [];
    shiftGroups[shiftType].push(card);
  }});

  function getCurrentShift(){{
    var now = new Date();
    var muscatTime = new Date(now.getTime() + (4 * 60 * 60 * 1000) + (now.getTimezoneOffset() * 60 * 1000));
    var t = muscatTime.getHours() * 60 + muscatTime.getMinutes();
    if(t >= 21 * 60 || t < 5 * 60) return 'Night';
    if(t >= 13 * 60) return 'Afternoon';
    return 'Morning';
  }}

  function setFilterBtnActive(selectedShift){{
    filterBtns.forEach(function(btn){{
      btn.classList.toggle('active', btnShift(btn) === selectedShift);
    }});
  }}

  function filterShifts(selectedShiftRaw){{
    var selectedShift = normalizeShift(selectedShiftRaw);
    window.__shiftFilterSelection = selectedShift;
    var totalEmployees = 0;

    if(selectedShift === 'All'){{
      allShiftCards.forEach(function(card){{
        card.style.display = '';
        var count = card.querySelector('.shiftCount');
        if(count) totalEmployees += parseInt(count.textContent, 10) || 0;
      }});
    }} else {{
      allShiftCards.forEach(function(card){{ card.style.display = 'none'; }});
      if(shiftGroups[selectedShift]){{
        shiftGroups[selectedShift].forEach(function(card){{
          card.style.display = '';
          card.setAttribute('open', '');
          var count = card.querySelector('.shiftCount');
          if(count) totalEmployees += parseInt(count.textContent, 10) || 0;
        }});
      }}
      var alwaysShow = ['Off Day', 'Annual Leave', 'Sick Leave', 'Training', 'Standby', 'Other'];
      alwaysShow.forEach(function(type){{
        if(shiftGroups[type]){{
          shiftGroups[type].forEach(function(card){{
            card.style.display = '';
          }});
        }}
      }});
    }}

    window.__summaryCounts = window.__summaryCounts || {{}};
    window.__summaryCounts.employees = totalEmployees;
    if(window.updateSummarySwitchChip) window.updateSummarySwitchChip();
    setFilterBtnActive(selectedShift);
  }}

  window.applyShiftFilter = filterShifts;

  filterBtns.forEach(function(btn){{
    btn.addEventListener('click', function(){{
      filterShifts(btnShift(btn));
    }});
  }});

  filterShifts(getCurrentShift());
}})();

// ═══════════════════════════════════════════════════
// رسالة الترحيب — تقرأ اسم الموظف من schedules JSON
// ═══════════════════════════════════════════════════
(function() {{
  function getExportEmpId() {{
    var id = localStorage.getItem('exportSavedEmpId');
    if (id) return id;
    var legacy = localStorage.getItem('savedEmpId');
    if (legacy) {{
      localStorage.setItem('exportSavedEmpId', legacy);
      return legacy;
    }}
    return '';
  }}
  var empId = getExportEmpId();
  if (!empId) return;
  var chip = document.getElementById('welcomeChip');
  var nameEl = document.getElementById('welcomeName');
  // استخدام مسار مطلق يعمل من أي صفحة
  var base = getSiteRootUrl() + '/';
  fetch(base + 'schedules/' + empId + '.json')
    .then(function(r) {{ return r.ok ? r.json() : null; }})
    .then(function(d) {{
      if (!d || !d.name) return;
      var firstName = d.name.split(' ')[0];
      if (chip && nameEl) {{
        nameEl.textContent = firstName;
        chip.classList.add('visible');
      }}
    }}).catch(function() {{}});
}})();

function goToMySchedule(event) {{
  if (event) event.preventDefault();
  var id = localStorage.getItem('exportSavedEmpId') || localStorage.getItem('savedEmpId');
  var base = getSiteRootUrl() + '/my-schedules/index.html';
  location.href = id ? base + '?emp=' + encodeURIComponent(id) : base;
}}

function goToImport(event) {{
  if (event) event.preventDefault();
  var picker = document.getElementById('datePicker');
  var iso = (picker && picker.value) ? picker.value : '';
  if (!iso) {{
    var m = (location.pathname || '').match(/\/date\/(\d{{4}}-\d{{2}}-\d{{2}})\//);
    if (m) iso = m[1];
  }}
  var root = getSiteRootUrl();
  var fallback = root + '/import/';
  if (!iso) {{
    location.href = fallback;
    return;
  }}
  fetch(root + '/import/import_meta.json', {{ cache: 'no-store' }})
    .then(function(r) {{ return r.ok ? r.json() : null; }})
    .then(function(meta) {{
      var months = meta && Array.isArray(meta.available_months) ? meta.available_months : [];
      var ym = iso.slice(0, 7);
      if (months.length && months.indexOf(ym) === -1) {{
        location.href = fallback;
        return;
      }}
      location.href = root + '/import/date/' + iso + '/';
    }})
    .catch(function() {{
      location.href = root + '/import/date/' + iso + '/';
    }});
}}

function goToRosterDiff(event) {{
  if (event) event.preventDefault();
  var target = getSiteRootUrl() + '/roster-diff/index.html';
  location.href = target;
}}

{LOAD_LOCAL_ENHANCEMENTS_EXPORT}
{EMPLOYEE_NEXT_SHIFT_PREVIEW_JS}
</script>

</body>
</html>"""


def generate_date_pages_for_month(
    wb,
    year: int,
    month: int,
    pages_base: str,
    source_name: str = "",
    min_date: str = "",
    max_date: str = "",
    site_last_updated: str = "",
):
    """
    Generate static pages for each day of the given month.
    Used by the date picker to navigate to different dates.

    If wb is None, it still generates pages but shows a 'no roster' notice.
    """
    import calendar
    from datetime import datetime as dt

    days_in_month = calendar.monthrange(year, month)[1]

    for day in range(1, days_in_month + 1):
        try:
            date_obj = dt(year, month, day, tzinfo=TZ)
            dow = (date_obj.weekday() + 1) % 7  # Sun=0
            active_group = current_shift_key(dt.now(TZ))

            dept_cards_all = []
            dept_cards_now = []
            employees_total_all = 0
            employees_total_now = 0
            depts_count = 0

            notice_html = ""
            if wb is None:
                notice_html = (
                    "<div class='deptCard' style='padding:14px;border:1px dashed rgba(15,23,42,.20);background:#fff;'>"
                    "⚠️ لا يوجد روستر لهذا الشهر بعد.</div>"
                )
            else:
                for idx, (sheet_name, dept_name) in enumerate(DEPARTMENTS):
                    if sheet_name not in wb.sheetnames:
                        continue

                    ws = wb[sheet_name]
                    days_row, date_row = find_days_and_dates_rows(ws)
                    day_col = find_day_col(ws, days_row, date_row, dow, day)

                    if not (days_row and date_row and day_col):
                        continue

                    start_row = date_row + 1
                    emp_col = find_employee_col(ws, start_row=start_row)
                    daynum_to_col = get_daynum_to_col(ws, date_row)
                    if not emp_col:
                        continue

                    buckets = {k: [] for k in GROUP_ORDER}
                    buckets_now = {k: [] for k in GROUP_ORDER}

                    for r in range(start_row, ws.max_row + 1):
                        name = norm(ws.cell(row=r, column=emp_col).value)
                        if not looks_like_employee_name(name):
                            continue

                        daynum_to_raw = {dn: norm(ws.cell(row=r, column=col).value) for dn, col in daynum_to_col.items()}
                        raw = daynum_to_raw.get(day, "")
                        if not looks_like_shift_code(raw):
                            continue

                        label, grp = map_shift(raw)
                        label = _apply_shift_range_label(label, grp, day, daynum_to_raw, raw)

                        buckets.setdefault(grp, []).append({"name": name, "shift": label})

                        # /now page: include ALL groups so the shift filter buttons work for any date
                        buckets_now.setdefault(grp, []).append({"name": name, "shift": label})
                    dept_color = UNASSIGNED_COLOR if dept_name == "Unassigned" else DEPT_COLORS[idx % len(DEPT_COLORS)]
                    open_group_full = active_group if AUTO_OPEN_ACTIVE_SHIFT_IN_FULL else None

                    dept_cards_all.append(dept_card_html(dept_name, dept_color, buckets, open_group=open_group_full))
                    dept_cards_now.append(dept_card_html(dept_name, dept_color, buckets_now, open_group=active_group))

                    employees_total_all += sum(len(buckets.get(g, [])) for g in GROUP_ORDER)
                    employees_total_now += sum(len(buckets_now.get(g, [])) for g in GROUP_ORDER)
                    depts_count += 1

                if employees_total_all == 0:
                    notice_html = (
                        "<div class='deptCard' style='padding:14px;border:1px dashed rgba(15,23,42,.20);background:#fff;'>"
                        "ℹ️ لا توجد بيانات لهذا التاريخ في الروستر.</div>"
                    )

            try:
                date_label = date_obj.strftime("%-d %B %Y")
            except Exception:
                date_label = date_obj.strftime("%d %B %Y")

            iso_date = date_obj.strftime("%Y-%m-%d")
            sent_time = date_obj.strftime("%H:%M")
            last_updated = site_last_updated

            full_url = f"{pages_base}/"
            now_url = f"{pages_base}/now/"

            html_full = page_shell_html(
                date_label=date_label,
                iso_date=iso_date,
                employees_total=employees_total_all,
                departments_total=depts_count,
                dept_cards_html="\n".join(dept_cards_all),
                cta_url=now_url,
                sent_time=sent_time,
                source_name=source_name,
                last_updated=last_updated,
                is_now_page=False,
                min_date=min_date,
                max_date=max_date,
                notice_html=notice_html,
            )

            html_now = page_shell_html(
                date_label=date_label,
                iso_date=iso_date,
                employees_total=employees_total_now,
                departments_total=depts_count,
                dept_cards_html="\n".join(dept_cards_now),
                cta_url=full_url,
                sent_time=sent_time,
                source_name=source_name,
                last_updated=last_updated,
                is_now_page=True,
                min_date=min_date,
                max_date=max_date,
                notice_html=notice_html,
            )

            date_dir = f"docs/date/{iso_date}"
            os.makedirs(date_dir, exist_ok=True)
            os.makedirs(f"{date_dir}/now", exist_ok=True)

            with open(f"{date_dir}/index.html", "w", encoding="utf-8") as f:
                f.write(html_full)

            with open(f"{date_dir}/now/index.html", "w", encoding="utf-8") as f:
                f.write(html_now)

        except Exception as e:
            print(f"Skipping {year}-{month:02d}-{day:02d}: {e}")
            continue


def build_pretty_email_html(active_shift_key: str, now: datetime, all_shifts_by_dept: list, pages_base: str) -> str:
    """
    Builds a beautifully formatted HTML email showing ONLY the active shift plus Standby for the same shift.
    all_shifts_by_dept = [{"dept": ..., "shifts": {"Morning": [...], "Afternoon": [...], ...}}, ...]
    """
    # Show only: active shift + Standby entries that match the active shift
    include_groups = [active_shift_key, "Standby"]

    def standby_matches_active(shift_text: str) -> bool:
        up = (shift_text or "").upper()
        if active_shift_key == "Morning":
            return bool(re.search(r"(MN|ME)\d{1,2}", up))
        if active_shift_key == "Afternoon":
            return bool(re.search(r"(AN|AE)\d{1,2}", up))
        if active_shift_key == "Night":
            return bool(re.search(r"(NN|NE|NT)\d{1,2}", up))
        return False

    def get_group_employees(shifts_data: dict, group_key: str):
        if group_key != "Standby":
            return shifts_data.get(group_key, []) or []
        # Standby: keep only those that match the active shift (e.g., STME06 for Morning)
        emps = shifts_data.get("Standby", []) or []
        return [e for e in emps if standby_matches_active(e.get("shift", ""))]

    # Calculate totals across included groups only
    total_employees = 0
    depts_with_employees = 0

    for d in all_shifts_by_dept:
        shifts_data = d.get("shifts", {})
        dept_total = 0
        for g in include_groups:
            dept_total += len(get_group_employees(shifts_data, g))
        if dept_total > 0:
            depts_with_employees += 1
            total_employees += dept_total

    # Determine current shift colors for header
    shift_colors = SHIFT_COLORS.get(active_shift_key, SHIFT_COLORS["Other"])
    shift_icon = shift_colors.get("icon", "⏰")

    # Build department cards with ALL shifts
    dept_cards = []
    for idx, d in enumerate(all_shifts_by_dept):
        dept_name = d["dept"]
        shifts_data = d["shifts"]
        # Skip if department has no employees for the included groups
        dept_total = 0
        for g in include_groups:
            dept_total += len(get_group_employees(shifts_data, g))
        if dept_total == 0:
            continue
        # Determine department color
        if dept_name == "Unassigned":
            dept_color = UNASSIGNED_COLOR
        else:
            dept_color = DEPT_COLORS[idx % len(DEPT_COLORS)]
        # Build shift sections (only active shift + matching Standby)
        shift_sections = ""
        for group_key in GROUP_ORDER:
            if group_key not in include_groups:
                continue
            employees = get_group_employees(shifts_data, group_key)
            if not employees:
                continue

            # Get shift display name
            shift_display_names = {
                "Morning": "Morning",
                "Afternoon": "Afternoon",
                "Night": "Night",
                "Off Day": "Off Day",
                "Leave": "Annual Leave",
                "Training": "Training",
                "Standby": "Standby",
                "Other": "Other"
            }
            display_name = shift_display_names.get(group_key, group_key)
            
            colors = SHIFT_COLORS.get(group_key, SHIFT_COLORS["Other"])
            count = len(employees)

            # Highlight active shift
            is_active = (group_key == active_shift_key)
            active_border = f"border:2px solid {colors['border']};" if is_active else f"border:1px solid {colors['border']};"
            active_badge = "⚡" if is_active else ""

            # Build employee rows
            rows_html = ""
            for i, e in enumerate(employees):
                bg_color = "rgba(15,23,42,.03)" if i % 2 == 1 else "transparent"
                rows_html += f"""
                    <tr>
                      <td style="padding:10px 14px;border-top:1px solid rgba(15,23,42,.06);background:{bg_color};">
                        <span style="font-size:14px;font-weight:700;color:#1e293b;">{e['name']}</span>
                      </td>
                      <td style="padding:10px 14px;border-top:1px solid rgba(15,23,42,.06);text-align:right;background:{bg_color};">
                        <span style="font-size:13px;font-weight:600;color:{colors['status_color']};white-space:nowrap;">{e['shift']}</span>
                      </td>
                    </tr>"""

            shift_sections += f"""
              <table role="presentation" cellpadding="0" cellspacing="0" style="width:100%;margin-top:10px;background:{colors['bg']};border-radius:12px;overflow:hidden;{active_border}">
                <!-- Shift Header -->
                <tr>
                  <td colspan="2" style="padding:10px 14px;background:{colors['summary_bg']};border-bottom:1px solid {colors['summary_border']};">
                    <table role="presentation" cellpadding="0" cellspacing="0" style="width:100%;">
                      <tr>
                        <td style="padding:0;">
                          <span style="font-size:18px;margin-right:8px;">{colors['icon']}</span>
                          <span style="font-size:15px;font-weight:800;color:{colors['label_color']};letter-spacing:-.1px;">{display_name} {active_badge}</span>
                        </td>
                        <td style="text-align:right;padding:0;">
                          <span style="display:inline-block;padding:4px 12px;border-radius:20px;background:{colors['count_bg']};color:{colors['count_color']};font-size:13px;font-weight:800;">{count}</span>
                        </td>
                      </tr>
                    </table>
                  </td>
                </tr>
                <!-- Employees -->
                {rows_html}
              </table>"""

        # Department icon SVG
        icon_svg = """<svg width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.2" stroke-linecap="round" stroke-linejoin="round">
  <path d="M3 21h18M3 10h18M5 21V10l7-6 7 6v11"/>
  <rect x="9" y="14" width="2" height="3"/>
  <rect x="13" y="14" width="2" height="3"/>
</svg>"""

        dept_cards.append(f"""
          <table role="presentation" cellpadding="0" cellspacing="0" style="width:100%;margin-top:18px;background:#fff;border-radius:18px;overflow:hidden;border:1px solid rgba(15,23,42,.07);box-shadow:0 4px 18px rgba(15,23,42,.08);">
            <!-- Colored top gradient bar -->
            <tr>
              <td colspan="2" style="height:5px;background:linear-gradient(to right,{dept_color['grad_from']},{dept_color['grad_to']});padding:0;"></td>
            </tr>
            
            <!-- Department Header -->
            <tr>
              <td colspan="2" style="padding:14px 16px;border-bottom:2px solid {dept_color['border']};background:#fff;">
                <table role="presentation" cellpadding="0" cellspacing="0" style="width:100%;">
                  <tr>
                    <td style="width:46px;padding:0;">
                      <div style="width:44px;height:44px;border-radius:12px;background:{dept_color['light']};color:{dept_color['base']};display:flex;align-items:center;justify-content:center;">
                        {icon_svg}
                      </div>
                    </td>
                    <td style="padding:0 0 0 12px;">
                      <span style="font-size:18px;font-weight:800;color:#1e293b;letter-spacing:-.2px;display:block;">{dept_name}</span>
                    </td>
                    <td style="text-align:right;padding:0;">
                      <div style="display:inline-block;min-width:52px;padding:8px 12px;border-radius:12px;background:{dept_color['light']};border:1px solid {dept_color['border']};text-align:center;">
                        <span style="font-size:10px;opacity:.7;display:block;text-transform:uppercase;letter-spacing:.5px;color:{dept_color['base']};margin-bottom:1px;">Total</span>
                        <span style="font-size:17px;font-weight:900;color:{dept_color['base']};display:block;">{dept_total}</span>
                      </div>
                    </td>
                  </tr>
                </table>
              </td>
            </tr>

            <!-- Included Shifts -->
            <tr>
              <td colspan="2" style="padding:10px;">
                {shift_sections}
              </td>
            </tr>
          </table>
        """)

    dept_html = "".join(dept_cards)
    sent_time = now.strftime("%H:%M")
    date_str = now.strftime("%d %B %Y")
    last_updated = format_site_last_updated(now)

    # Translate active_shift_key display
    shift_display_map = {
        "Morning": "Morning Shift",
        "Afternoon": "Afternoon Shift", 
        "Night": "Night Shift"
    }
    shift_display = shift_display_map.get(active_shift_key, active_shift_key)

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <meta name="x-apple-disable-message-reformatting">
  <title>Duty Roster - {date_str}</title>
  <style>
    @media only screen and (max-width: 600px) {{
      .mobile-padding {{ padding: 12px !important; }}
      .mobile-font {{ font-size: 13px !important; }}
      .header-icon {{ font-size: 56px !important; }}
    }}
  </style>
</head>
<body style="margin:0;padding:0;background:#eef1f7;font-family:'Segoe UI',system-ui,-apple-system,BlinkMacSystemFont,Roboto,Helvetica,Arial,sans-serif;-webkit-font-smoothing:antialiased;">
  
  <table role="presentation" cellpadding="0" cellspacing="0" border="0" style="width:100%;background:#eef1f7;">
    <tr>
      <td align="center" style="padding:20px 14px;">
        
        <!-- Main Container -->
        <table role="presentation" cellpadding="0" cellspacing="0" border="0" style="max-width:680px;width:100%;margin:0 auto;">
          
          <!-- Large Header with Gradient -->
          <tr>
<!-- Compact Header with Gradient -->
<tr>
  <td style="padding:0;">
    <table role="presentation" cellpadding="0" cellspacing="0" border="0"
      style="width:100%;
             background:linear-gradient(135deg,#1e40af 0%,#1976d2 50%,#0ea5e9 100%);
             border-radius:20px 20px 0 0;
             overflow:hidden;
             box-shadow:0 8px 26px rgba(30,64,175,.25);
             position:relative;">
      <tr>
        <td style="padding:18px 18px;text-align:center;position:relative;">

          <!-- Decorative circles -->
          <div style="position:absolute;top:-50px;right:-60px;width:160px;height:160px;border-radius:50%;background:rgba(255,255,255,.08);"></div>
          <div style="position:absolute;bottom:-70px;left:-50px;width:180px;height:180px;border-radius:50%;background:rgba(255,255,255,.06);"></div>

          <!-- Icon -->
          <div class="header-icon"
               style="font-size:40px;margin-bottom:6px;position:relative;z-index:1;">📋</div>

          <!-- Title -->
          <h1 style="margin:0;
                     font-size:22px;
                     font-weight:800;
                     color:#ffffff;
                     letter-spacing:-.4px;
                     position:relative;
                     z-index:1;">
            Duty Roster
          </h1>

          <!-- Active Shift -->
          <div style="margin-top:8px;
                      display:inline-block;
                      background:rgba(255,255,255,.22);
                      padding:6px 16px;
                      border-radius:18px;
                      font-size:13px;
                      font-weight:700;
                      color:#ffffff;
                      letter-spacing:.3px;
                      position:relative;
                      z-index:1;">
            {shift_icon} {shift_display}
          </div>

          <!-- Date -->
          <div style="margin-top:6px;
                      display:inline-block;
                      background:rgba(255,255,255,.16);
                      padding:5px 14px;
                      border-radius:16px;
                      font-size:12px;
                      font-weight:600;
                      color:#ffffff;
                      letter-spacing:.2px;
                      position:relative;
                      z-index:1;">
            📅 {date_str}
          </div>

        </td>
      </tr>
    </table>
  </td>
</tr>

          <!-- Summary Stats -->
          <tr>
            <td style="padding:0 14px;">
              <table role="presentation" cellpadding="0" cellspacing="0" border="0" style="width:100%;margin-top:18px;">
                <tr>
                  <td style="width:50%;padding-right:6px;">
                    <table role="presentation" cellpadding="0" cellspacing="0" border="0" style="width:100%;background:#fff;border:1px solid rgba(15,23,42,.10);border-radius:16px;box-shadow:0 3px 12px rgba(15,23,42,.07);">
                      <tr>
                        <td style="padding:16px;text-align:center;">
                          <div style="font-size:28px;font-weight:900;color:#1e40af;margin-bottom:4px;">{total_employees}</div>
                          <div style="font-size:12px;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:.6px;">Employees</div>
                        </td>
                      </tr>
                    </table>
                  </td>
                  <td style="width:50%;padding-left:6px;">
                    <table role="presentation" cellpadding="0" cellspacing="0" border="0" style="width:100%;background:#fff;border:1px solid rgba(15,23,42,.10);border-radius:16px;box-shadow:0 3px 12px rgba(15,23,42,.07);">
                      <tr>
                        <td style="padding:16px;text-align:center;">
                          <div style="font-size:28px;font-weight:900;color:#059669;margin-bottom:4px;">{depts_with_employees}</div>
                          <div style="font-size:12px;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:.6px;">Departments</div>
                        </td>
                      </tr>
                    </table>
                  </td>
                </tr>
              </table>
            </td>
          </tr>

          <!-- Department Cards with ALL Shifts -->
          <tr>
            <td style="padding:0 14px;">
              {dept_html}
            </td>
          </tr>

          <!-- Call to Action Buttons -->
          <tr>
            <td style="padding:22px 14px;text-align:center;">
              <table role="presentation" cellpadding="0" cellspacing="0" border="0" style="margin:0 auto;">
                <tr>
                  <td style="padding:0 7px 0 0;">
                    <a href="{pages_base}/now/" style="display:inline-block;padding:15px 30px;border-radius:16px;background:linear-gradient(135deg,#1e40af,#1976d2);color:#ffffff;text-decoration:none;font-weight:800;font-size:15px;box-shadow:0 6px 22px rgba(30,64,175,.35);white-space:nowrap;">
                      🔄 Refresh Now
                    </a>
                  </td>
                  <td style="padding:0 0 0 7px;">
                    <a href="{pages_base}/" style="display:inline-block;padding:15px 30px;border-radius:16px;background:linear-gradient(135deg,#0ea5e9,#06b6d4);color:#ffffff;text-decoration:none;font-weight:800;font-size:15px;box-shadow:0 6px 22px rgba(14,165,233,.35);white-space:nowrap;">
                      📋 View Full Roster
                    </a>
                  </td>
                </tr>
              </table>
            </td>
          </tr>

          <!-- Footer -->
          <tr>
            <td style="padding:0 14px 22px;">
              <table role="presentation" cellpadding="0" cellspacing="0" border="0" style="width:100%;background:#fff;border-radius:0 0 20px 20px;border:1px solid rgba(15,23,42,.08);border-top:none;">
                <tr>
                  <td style="padding:18px;text-align:center;color:#94a3b8;font-size:13px;line-height:1.9;">
                    <strong style="color:#475569;">Last Updated:</strong> <strong style="color:#1e40af;">{last_updated}</strong>
                    <br>
                    Total on duty: <strong style="color:#64748b;">{total_employees} employees</strong> across <strong style="color:#64748b;">{depts_with_employees} departments</strong>
                  </td>
                </tr>
              </table>
            </td>
          </tr>

        </table>
        
      </td>
    </tr>
  </table>

</body>
</html>"""




# =========================
# Main
# =========================
def main():
    parser = argparse.ArgumentParser(description='Generate roster pages and send email')
    parser.add_argument('--date', help='Override roster date (YYYY-MM-DD)')
    parser.add_argument('--no-email', action='store_true', help='Generate pages only, do not send email')
    parser.add_argument('--excel-file', help='Use local Excel file instead of EXCEL_URL download')
    parser.add_argument('--source-name', help='Optional source filename to display and month-detect when using --excel-file')
    args = parser.parse_args()

    now = datetime.now(TZ)
    if args.date:
        try:
            y, m, d = [int(x) for x in args.date.strip().split('-')]
            now = datetime(y, m, d, now.hour, now.minute, tzinfo=TZ)
        except Exception:
            raise RuntimeError('Invalid --date format. Use YYYY-MM-DD')

    today_dow = (now.weekday() + 1) % 7
    today_day = now.day
    active_group = current_shift_key(now)

    # pages_base - cleanup
    pages_base_raw = PAGES_BASE_URL or infer_pages_base_url()
    pages_base = pages_base_raw.rstrip("/")
    if pages_base.endswith("/now"):
        pages_base = pages_base[:-4]

    # ─────────────────────────────────────────────────────────────
    # FIX #1: تحميل Excel - عند الفشل نكمل بالكاش (لا نخرج)
    # ─────────────────────────────────────────────────────────────
    data = None
    if args.excel_file:
        with open(args.excel_file, "rb") as f:
            data = f.read()
        print(f"Using local Excel file: {args.excel_file}")
    elif EXCEL_URL:
        try:
            data = download_excel(EXCEL_URL)
            print("✅ Excel downloaded successfully")
        except Exception as e:
            print(f"WARNING: Could not download Excel: {e}")
            print("Will attempt to use cached rosters...")
    else:
        print("⚠️ EXCEL_URL missing; using cached rosters only.")

    # ─────────────────────────────────────────────────────────────
    # FIX #2: كل هذا الكود الآن خارج الـ except - يعمل دائماً
    # ─────────────────────────────────────────────────────────────

    # قراءة اسم الملف واستخراج الشهر
    source_name = (args.source_name or "").strip() or get_source_name()
    incoming_key = month_key_from_filename(source_name) if source_name else None
    print(f"📄 Source file: {source_name}")
    print(f"📅 Detected month: {incoming_key or 'unknown'}")
    if source_name and looks_like_roster_month_filename(source_name) and not incoming_key:
        raise RuntimeError(
            f"Could not detect month from roster filename: {source_name}. "
            "Fix month_key_from_filename before publishing."
        )

    # Anchor calendar window to the roster file month (e.g. June file while today is still May).
    if incoming_key and not args.date and (args.excel_file or data):
        try:
            ay, am = [int(x) for x in incoming_key.split("-")]
            if (ay, am) != (now.year, now.month):
                now = datetime(ay, am, 1, now.hour, now.minute, tzinfo=TZ)
                today_dow = (now.weekday() + 1) % 7
                today_day = now.day
                active_group = current_shift_key(now)
                print(f"📅 Anchored publish month to file: {incoming_key}")
        except Exception:
            pass

    site_last_updated = format_site_last_updated(now)
    write_site_last_updated_json(now)

    # FIX #3: حفظ في الكاش فقط إذا نجح التحميل
    if data and incoming_key:
        xlsx_path, meta_path = cache_paths(incoming_key)
        try:
            write_bytes(xlsx_path, data)
            write_json(meta_path, {
                "month_key": incoming_key,
                "original_filename": source_name,
                "downloaded_at": datetime.now(TZ).strftime("%Y-%m-%d %H:%M:%S %Z"),
            })
            print(f"✅ Cached roster: {xlsx_path}")
        except Exception as e:
            print(f"WARNING: failed caching roster: {e}")
    elif not incoming_key:
        print("⚠️ Could not detect month from filename; cache skipped for this run.")

    # حساب الأشهر الثلاثة
    prev_y, prev_m = add_months(now.year, now.month, -1)
    next_y, next_m = add_months(now.year, now.month, +1)

    prev_key = f"{prev_y:04d}-{prev_m:02d}"
    curr_key = f"{now.year:04d}-{now.month:02d}"
    next_key = f"{next_y:04d}-{next_m:02d}"

    # نطاق الـ date picker: من أول الشهر السابق إلى آخر الشهر القادم
    min_date = f"{prev_y:04d}-{prev_m:02d}-01"
    max_date = f"{next_y:04d}-{next_m:02d}-{calendar.monthrange(next_y, next_m)[1]:02d}"

    print(f"📅 Month range: {prev_key} → {curr_key} → {next_key}")

    # تحميل الكاش لكل شهر
    wb_prev = try_load_cached_workbook(prev_key)
    wb_curr = try_load_cached_workbook(curr_key)
    wb_next = try_load_cached_workbook(next_key)

    # FIX #4: استخدام البيانات المحملة للشهر المطابق (تجاوز الكاش القديم)
    if data:
        wb_data = load_workbook(BytesIO(data), data_only=True)
        if incoming_key == prev_key:
            wb_prev = wb_data
            print(f"✅ Using downloaded data for {prev_key}")
        if incoming_key == curr_key:
            wb_curr = wb_data
            print(f"✅ Using downloaded data for {curr_key}")
        if incoming_key == next_key:
            wb_next = wb_data
            print(f"✅ Using downloaded data for {next_key}")

    print(f"📦 Cache status: prev={'✅' if wb_prev else '❌'} | curr={'✅' if wb_curr else '❌'} | next={'✅' if wb_next else '❌'}")

    # توليد صفحات الأشهر الثلاثة
    generate_date_pages_for_month(
        wb_prev, prev_y, prev_m, pages_base,
        source_name=cached_source_name(prev_key) or source_name,
        min_date=min_date, max_date=max_date,
        site_last_updated=site_last_updated,
    )
    generate_date_pages_for_month(
        wb_curr, now.year, now.month, pages_base,
        source_name=cached_source_name(curr_key) or source_name,
        min_date=min_date, max_date=max_date,
        site_last_updated=site_last_updated,
    )
    generate_date_pages_for_month(
        wb_next, next_y, next_m, pages_base,
        source_name=cached_source_name(next_key) or source_name,
        min_date=min_date, max_date=max_date,
        site_last_updated=site_last_updated,
    )

    if incoming_key and data:
        slot_wb = {prev_key: wb_prev, curr_key: wb_curr, next_key: wb_next}.get(incoming_key)
        if slot_wb is None:
            raise RuntimeError(
                f"Roster data downloaded for {incoming_key} but no workbook was bound "
                f"to the publish window ({prev_key}, {curr_key}, {next_key})."
            )

    # الصفحة الرئيسية تستخدم الشهر الحالي
    wb = wb_curr

    # ─────────────────────────────────────────────────────────────
    # من هنا: توليد الصفحة الرئيسية docs/index.html و docs/now/
    # ─────────────────────────────────────────────────────────────
    if wb is None:
        os.makedirs("docs", exist_ok=True)
        os.makedirs("docs/now", exist_ok=True)

        try:
            date_label = now.strftime("%-d %B %Y")
        except Exception:
            date_label = now.strftime("%d %B %Y")

        iso_date = now.strftime("%Y-%m-%d")
        last_updated = site_last_updated

        notice_html = (
            "<div class='deptCard' style='padding:14px;border:1px dashed rgba(15,23,42,.20);background:#fff;'>"
            "⚠️ لا يوجد روستر للشهر الحالي محفوظ بعد. (قد يكون الروستر الجديد للشهر القادم وصل مبكرًا)</div>"
        )

        html_full = page_shell_html(
            date_label=date_label,
            iso_date=iso_date,
            employees_total=0,
            departments_total=0,
            dept_cards_html="",
            cta_url=f"{pages_base}/now/",
            sent_time=now.strftime("%H:%M"),
            source_name=cached_source_name(curr_key) or source_name,
            last_updated=last_updated,
            is_now_page=False,
            min_date=min_date,
            max_date=max_date,
            notice_html=notice_html,
        )

        html_now = page_shell_html(
            date_label=date_label,
            iso_date=iso_date,
            employees_total=0,
            departments_total=0,
            dept_cards_html="",
            cta_url=f"{pages_base}/",
            sent_time=now.strftime("%H:%M"),
            source_name=cached_source_name(curr_key) or source_name,
            last_updated=last_updated,
            is_now_page=True,
            min_date=min_date,
            max_date=max_date,
            notice_html=notice_html,
        )

        with open("docs/index.html", "w", encoding="utf-8") as f:
            f.write(html_full)

        with open("docs/now/index.html", "w", encoding="utf-8") as f:
            f.write(html_now)

        print("⚠️ Skipping email (no current-month roster workbook).")
        return


    dept_cards_all = []
    dept_cards_now = []
    all_shifts_by_dept = []
    employees_total_all = 0
    employees_total_now = 0
    depts_count = 0

    for idx, (sheet_name, dept_name) in enumerate(DEPARTMENTS):
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        days_row, date_row = find_days_and_dates_rows(ws)
        day_col = find_day_col(ws, days_row, date_row, today_dow, today_day)

        if not (days_row and date_row and day_col):
            continue

        start_row = date_row + 1
        emp_col = find_employee_col(ws, start_row=start_row)
        daynum_to_col = get_daynum_to_col(ws, date_row)
        if not emp_col:
            continue

        buckets = {k: [] for k in GROUP_ORDER}
        buckets_now = {k: [] for k in GROUP_ORDER}

        for r in range(start_row, ws.max_row + 1):
            name = norm(ws.cell(row=r, column=emp_col).value)
            if not looks_like_employee_name(name):
                continue

            daynum_to_raw = {dn: norm(ws.cell(row=r, column=col).value) for dn, col in daynum_to_col.items()}

            raw = daynum_to_raw.get(today_day, "")
            if not looks_like_shift_code(raw):
                continue

            label, grp = map_shift(raw)
            label = _apply_shift_range_label(label, grp, today_day, daynum_to_raw, raw)

            buckets.setdefault(grp, []).append({"name": name, "shift": label})

            if grp == active_group:
                buckets_now.setdefault(grp, []).append({"name": name, "shift": label})

        all_shifts_by_dept.append({"dept": dept_name, "shifts": buckets})

        if dept_name == "Unassigned":
            dept_color = UNASSIGNED_COLOR
        else:
            dept_color = DEPT_COLORS[idx % len(DEPT_COLORS)]

        open_group_full = active_group if AUTO_OPEN_ACTIVE_SHIFT_IN_FULL else None
        card_all = dept_card_html(dept_name, dept_color, buckets, open_group=open_group_full)
        dept_cards_all.append(card_all)

        # صفحة /now/ تحتوي على كل الورديات (سيتم الفلترة بـ JavaScript)
        card_now = dept_card_html(dept_name, dept_color, buckets, open_group=active_group)
        dept_cards_now.append(card_now)

        employees_total_all += sum(len(buckets.get(g, [])) for g in GROUP_ORDER)
        # حساب فقط الوردية الحالية لـ total في /now/
        employees_total_now += sum(len(buckets.get(g, [])) for g in [active_group, "Off Day", "Annual Leave", "Sick Leave", "Training", "Standby", "Other"])

        depts_count += 1

    os.makedirs("docs", exist_ok=True)
    os.makedirs("docs/now", exist_ok=True)

    try:
        date_label = now.strftime("%-d %B %Y")
    except Exception:
        date_label = now.strftime("%d %B %Y")

    iso_date = now.strftime("%Y-%m-%d")
    sent_time = now.strftime("%H:%M")
    last_updated = site_last_updated

    full_url = f"{pages_base}/"
    now_url = f"{pages_base}/now/"

    html_full = page_shell_html(
        date_label=date_label,
        iso_date=iso_date,
        employees_total=employees_total_all,
        departments_total=depts_count,
        dept_cards_html="\n".join(dept_cards_all),
        cta_url=now_url,
        sent_time=sent_time,
        source_name=cached_source_name(curr_key) or source_name,
        last_updated=last_updated,
        min_date=min_date,
        max_date=max_date,
        is_now_page=False,
    )
    html_now = page_shell_html(
        date_label=date_label,
        iso_date=iso_date,
        employees_total=employees_total_now,
        departments_total=depts_count,
        dept_cards_html="\n".join(dept_cards_now),
        cta_url=full_url,
        sent_time=sent_time,
        source_name=cached_source_name(curr_key) or source_name,
        last_updated=last_updated,
        min_date=min_date,
        max_date=max_date,
        is_now_page=True,
    )

    with open("docs/index.html", "w", encoding="utf-8") as f:
        f.write(html_full)

    with open("docs/now/index.html", "w", encoding="utf-8") as f:
        f.write(html_now)

    # Write source name for my-schedules page to display
    _src = cached_source_name(curr_key) or source_name
    if _src:
        os.makedirs("docs/my-schedules", exist_ok=True)
        with open("docs/my-schedules/source.txt", "w", encoding="utf-8") as f:
            f.write(_src)

    write_site_last_updated_json(datetime.now(TZ))

    # Persist any newly discovered Arabic name translations for owner review.
    try:
        name_i18n.flush()
    except Exception as e:
        print(f"WARNING: could not write name translations: {e}")

    if source_name:
        try:
            with open("last_filename.txt", "w", encoding="utf-8") as f:
                f.write(source_name.strip())
        except Exception as e:
            print(f"WARNING: could not write last_filename.txt: {e}")

    # Email: send ONLY active shift + matching Standby
    if args.no_email:
        print("ℹ️ Email disabled via --no-email")
    else:
        subject = f"Duty Roster — {now.strftime('%d %B %Y')} — {active_group} Active"
        email_html = build_pretty_email_html(active_group, now, all_shifts_by_dept, pages_base)
        send_email(subject, email_html)


if __name__ == "__main__":
    main()